# -*- coding: utf-8 -*-
"""
内容ベース比較（DiffEngine）用サンプル .xlsx を生成する。

出力先: 親ディレクトリ (30_参考資料/samples/)
  - content_diff_left.xlsx
  - content_diff_right.xlsx

中間メディア: _gen/media_content_diff/

シート設計（設計書 §7.1 シナリオ）:

| シート名       | 左 | 右 | 期待 |
|----------------|----|----|------|
| S_Cells        | A1=Hello | A2=Hello | 位置無視 → Text 差なし |
| S_Bg           | A1=Hello 赤 | B2=Hello 白 | Background |
| S_TableDel     | 行 1..5 表 | 行 1,2,4,5 表 | TableRowDelete×1（3） |
| S_TableCell    | Hello/World | Hello/Changed | TableCellChange×1 |
| S_ImgSame      | 同見た目画像 @B5 | 同見た目画像 @D20 | Image 差なし |
| S_Img8v9       | 画像 8 枚 | 画像 9 枚（5 枚目が余分） | ImageOnlyRight×1 |
| S_ImgPartial   | base 画像 | 部分スタンプ画像 | Image + Regions≥1 |
| S_Common       | 同一内容 | 同一内容 | 差分なし |
| S_LeftOnly     | 左のみシート | （無し） | Structure |

再生成:
  python 30_参考資料/samples/_gen/create_content_diff_samples.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT.parent
MEDIA = ROOT / "media_content_diff"
MEDIA.mkdir(parents=True, exist_ok=True)

THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
BODY = Font(name="Yu Gothic UI", size=11)
TITLE = Font(name="Yu Gothic UI", size=12, bold=True, color="1F4E79")
RED_FILL = PatternFill("solid", fgColor="FF0000")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="FFFFFF")


def _font(size: int = 22):
    for name in ("segoeui.ttf", "arial.ttf", "YuGothR.ttc"):
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def make_solid(path: Path, label: str, bg: tuple, size=(160, 100), accent=None):
    """ラベル付きソリッド PNG。"""
    img = Image.new("RGB", size, bg)
    draw = ImageDraw.Draw(img)
    if accent is None:
        accent = (255, 255, 255)
    draw.rectangle([4, 4, size[0] - 5, size[1] - 5], outline=accent, width=3)
    font = _font(20)
    bbox = draw.textbbox((0, 0), label, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text(((size[0] - tw) / 2, (size[1] - th) / 2), label, fill=accent, font=font)
    img.save(path, "PNG")
    return path


def make_partial_mod(src: Path, dst: Path, stamp: str = "MOD"):
    """右下に赤い矩形スタンプを重ねる（部分差）。"""
    img = Image.open(src).convert("RGB")
    draw = ImageDraw.Draw(img)
    w, h = img.size
    draw.rectangle([w // 2, h // 2, w - 8, h - 8], fill=(220, 40, 40))
    font = _font(18)
    bbox = draw.textbbox((0, 0), stamp, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text(
        (w // 2 + (w // 2 - 8 - tw) / 2, h // 2 + (h // 2 - 8 - th) / 2),
        stamp,
        fill=(255, 255, 255),
        font=font,
    )
    img.save(dst, "PNG")
    return dst


def bordered_cell(ws, row, col, value, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN
    cell.font = BODY
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill is not None:
        cell.fill = fill
    return cell


def add_table(ws, start_row: int, rows: list[list[str]], header: list[str] | None = None):
    """ボーダー付き 2 列以上の表を書く（TableDetector は min 2x2）。"""
    r = start_row
    if header:
        for c, h in enumerate(header, start=1):
            cell = bordered_cell(ws, r, c, h, HEADER_FILL)
            cell.font = HEADER_FONT
        r += 1
    for row_vals in rows:
        for c, v in enumerate(row_vals, start=1):
            bordered_cell(ws, r, c, v)
        r += 1
    return r


def place_image(ws, path: Path, anchor: str):
    img = XLImage(str(path))
    img.anchor = anchor
    ws.add_image(img)


def build_media():
    """シナリオ用 PNG を用意する。"""
    same = make_solid(MEDIA / "same_visual.png", "SAME", (40, 120, 200))
    base = make_solid(MEDIA / "partial_base.png", "BASE", (60, 60, 90), size=(200, 140))
    partial = make_partial_mod(base, MEDIA / "partial_mod.png", "MOD")

    seq = []
    colors = [
        (30, 100, 180),
        (40, 140, 90),
        (160, 80, 40),
        (120, 50, 140),
        (40, 140, 160),
        (180, 120, 40),
        (80, 80, 80),
        (20, 60, 120),
    ]
    for i, bg in enumerate(colors):
        p = make_solid(MEDIA / f"seq_{i}.png", f"S{i}", bg, size=(120, 80))
        seq.append(p)
    insert = make_solid(MEDIA / "seq_insert.png", "INS", (200, 40, 40), size=(120, 80))
    return {
        "same": same,
        "base": base,
        "partial": partial,
        "seq": seq,
        "insert": insert,
    }


def build_left(media: dict) -> Workbook:
    wb = Workbook()
    # --- S_Cells ---
    ws = wb.active
    ws.title = "S_Cells"
    ws["A1"] = "Hello"
    ws["A1"].font = BODY
    ws["C1"] = "position-agnostic Hello"
    ws["C1"].font = TITLE

    # --- S_Bg ---
    ws = wb.create_sheet("S_Bg")
    ws["A1"] = "Hello"
    ws["A1"].font = BODY
    ws["A1"].fill = RED_FILL
    ws["C1"] = "bg red vs white"
    ws["C1"].font = TITLE

    # --- S_TableDel ---
    ws = wb.create_sheet("S_TableDel")
    ws["A1"] = "table 12345 vs 1245"
    ws["A1"].font = TITLE
    add_table(
        ws,
        3,
        [[str(i), f"row{i}"] for i in (1, 2, 3, 4, 5)],
        header=["Id", "Label"],
    )

    # --- S_TableCell ---
    ws = wb.create_sheet("S_TableCell")
    ws["A1"] = "table cell change"
    ws["A1"].font = TITLE
    add_table(
        ws,
        3,
        [["Hello", "World"], ["Keep", "Same"]],
        header=["ColA", "ColB"],
    )

    # --- S_ImgSame ---
    ws = wb.create_sheet("S_ImgSame")
    ws["A1"] = "same visual different position"
    ws["A1"].font = TITLE
    place_image(ws, media["same"], "B5")

    # --- S_Img8v9 ---
    ws = wb.create_sheet("S_Img8v9")
    ws["A1"] = "images 8 vs 9 (insert after 4th)"
    ws["A1"].font = TITLE
    # 8 images in reading order rows 3,5,7,...
    for i, p in enumerate(media["seq"]):
        place_image(ws, p, f"B{3 + i * 2}")

    # --- S_ImgPartial ---
    ws = wb.create_sheet("S_ImgPartial")
    ws["A1"] = "partial visual regions"
    ws["A1"].font = TITLE
    place_image(ws, media["base"], "B3")

    # --- S_Common ---
    ws = wb.create_sheet("S_Common")
    ws["A1"] = "Shared"
    ws["A1"].font = BODY
    ws["B2"] = 42

    # --- S_LeftOnly ---
    ws = wb.create_sheet("S_LeftOnly")
    ws["A1"] = "left only sheet"
    ws["A1"].font = BODY

    return wb


def build_right(media: dict) -> Workbook:
    wb = Workbook()
    # --- S_Cells: Hello at A2 (different position) ---
    ws = wb.active
    ws.title = "S_Cells"
    ws["A2"] = "Hello"
    ws["A2"].font = BODY
    ws["C1"] = "position-agnostic Hello"
    ws["C1"].font = TITLE

    # --- S_Bg: Hello white at B2 ---
    ws = wb.create_sheet("S_Bg")
    ws["B2"] = "Hello"
    ws["B2"].font = BODY
    ws["B2"].fill = WHITE_FILL
    ws["C1"] = "bg red vs white"
    ws["C1"].font = TITLE

    # --- S_TableDel: without row 3 ---
    ws = wb.create_sheet("S_TableDel")
    ws["A1"] = "table 12345 vs 1245"
    ws["A1"].font = TITLE
    add_table(
        ws,
        3,
        [[str(i), f"row{i}"] for i in (1, 2, 4, 5)],
        header=["Id", "Label"],
    )

    # --- S_TableCell ---
    ws = wb.create_sheet("S_TableCell")
    ws["A1"] = "table cell change"
    ws["A1"].font = TITLE
    add_table(
        ws,
        3,
        [["Hello", "Changed"], ["Keep", "Same"]],
        header=["ColA", "ColB"],
    )

    # --- S_ImgSame: same image different anchor ---
    ws = wb.create_sheet("S_ImgSame")
    ws["A1"] = "same visual different position"
    ws["A1"].font = TITLE
    place_image(ws, media["same"], "D20")

    # --- S_Img8v9: 9 images, insert after first 4 ---
    ws = wb.create_sheet("S_Img8v9")
    ws["A1"] = "images 8 vs 9 (insert after 4th)"
    ws["A1"].font = TITLE
    row = 3
    for i in range(4):
        place_image(ws, media["seq"][i], f"B{row}")
        row += 2
    place_image(ws, media["insert"], f"B{row}")
    row += 2
    for i in range(4, 8):
        place_image(ws, media["seq"][i], f"B{row}")
        row += 2

    # --- S_ImgPartial ---
    ws = wb.create_sheet("S_ImgPartial")
    ws["A1"] = "partial visual regions"
    ws["A1"].font = TITLE
    place_image(ws, media["partial"], "B3")

    # --- S_Common ---
    ws = wb.create_sheet("S_Common")
    ws["A1"] = "Shared"
    ws["A1"].font = BODY
    ws["B2"] = 42

    # no S_LeftOnly on right

    return wb


def main():
    media = build_media()
    left = build_left(media)
    right = build_right(media)
    left_path = OUT_DIR / "content_diff_left.xlsx"
    right_path = OUT_DIR / "content_diff_right.xlsx"
    left.save(left_path)
    right.save(right_path)
    print("Wrote", left_path)
    print("Wrote", right_path)
    print("Media dir:", MEDIA)


if __name__ == "__main__":
    main()
