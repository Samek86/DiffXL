# -*- coding: utf-8 -*-
"""
大きな埋め込み画像を含む DiffXL 検証用 .xlsx を生成する。

出力:
  - large_image_left.xlsx
  - large_image_right.xlsx

検証観点:
  - 高解像度 PNG / JPEG の抽出・OpenCV 比較
  - 同一大画像 / 内容差分 / 片側のみ
  - 複数大画像が同一シートにある場合の対応付け
  - MiniMap・テキスト差分（full_feature 互換シート名で auto-live も可）
  - ファイルサイズ・比較時間のストレス
"""

from __future__ import annotations

import hashlib
import os
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT.parent
MEDIA = ROOT / "media_large"
MEDIA.mkdir(parents=True, exist_ok=True)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Yu Gothic UI", size=16, bold=True, color="1F4E79")
SUB_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="2E75B6")
BODY_FONT = Font(name="Yu Gothic UI", size=11)
MONO_FONT = Font(name="Consolas", size=10)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
DIFF_HINT_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
ALT_ROW = PatternFill("solid", fgColor="D6EAF8")


def _font(size: int):
    for name in ("segoeui.ttf", "arial.ttf", "YuGothM.ttc", "meiryo.ttc"):
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def make_large_image(
    path: Path,
    size: tuple[int, int],
    label: str,
    bg: tuple[int, int, int],
    accent: tuple[int, int, int],
    fmt: str = "PNG",
    quality: int = 92,
    noise_seed: int = 1,
    stamp: str | None = None,
    bulk: bool = True,
) -> Path:
    """
    大きなテスト画像を生成する。
    bulk=True のとき圧縮しにくいノイズを敷き、実ファイルサイズを大きくする。
    noise_seed を変えると画素パターンが変わり、同一ラベルでも内容差分になる。
    stamp を指定すると右下に大きな差分スタンプを描く。
    """
    import numpy as np

    w, h = size
    # ベース: グラデーション + 決定論的ノイズ（ファイルを肥大化）
    rng = np.random.RandomState((noise_seed * 7919 + w + h) & 0xFFFFFFFF)
    if bulk:
        # RGB ノイズをベース色に混ぜて PNG/JPEG が小さくなりすぎないようにする
        noise = rng.randint(0, 256, (h, w, 3), dtype=np.uint8)
        yy = np.linspace(0, 1, h, dtype=np.float32).reshape(h, 1, 1)
        base = np.array(bg, dtype=np.float32).reshape(1, 1, 3)
        acc = np.array(accent, dtype=np.float32).reshape(1, 1, 3)
        grad = base * (1.0 - yy) + acc * yy
        mix = (grad * 0.45 + noise.astype(np.float32) * 0.55).clip(0, 255).astype(np.uint8)
        img = Image.fromarray(mix, mode="RGB")
    else:
        img = Image.new("RGB", (w, h), bg)

    draw = ImageDraw.Draw(img)

    # グリッド（視覚的アンカー）
    step = max(40, min(w, h) // 20)
    grid = tuple(max(0, min(255, c + 40)) for c in accent)
    for x in range(0, w, step):
        draw.line([(x, 0), (x, h)], fill=grid, width=1)
    for y in range(0, h, step):
        draw.line([(0, y), (w, y)], fill=grid, width=1)

    # 枠・ヘッダ
    draw.rectangle([12, 12, w - 13, h - 13], outline=accent, width=8)
    draw.rectangle([12, 12, w - 13, 90], fill=accent)

    font = _font(48)
    bbox = draw.textbbox((0, 0), label, font=font)
    tw = bbox[2] - bbox[0]
    draw.text(((w - tw) / 2, 24), label, fill=(255, 255, 255), font=font)

    meta = f"{w}x{h} {fmt} seed={noise_seed}"
    f2 = _font(28)
    draw.text((24, 110), meta, fill=(255, 255, 255), font=f2)

    if stamp:
        sw, sh = w // 3, h // 4
        x0, y0 = w - sw - 40, h - sh - 40
        draw.rectangle([x0, y0, x0 + sw, y0 + sh], fill=(220, 40, 40))
        fs = _font(max(36, sh // 3))
        bbox3 = draw.textbbox((0, 0), stamp, font=fs)
        tw3, th3 = bbox3[2] - bbox3[0], bbox3[3] - bbox3[1]
        draw.text(
            (x0 + (sw - tw3) / 2, y0 + (sh - th3) / 2),
            stamp,
            fill=(255, 255, 255),
            font=fs,
        )
        draw.polygon([(40, 120), (200, 120), (40, 280)], fill=(255, 210, 0))

    path.parent.mkdir(parents=True, exist_ok=True)
    if fmt.upper() == "JPEG" or path.suffix.lower() in (".jpg", ".jpeg"):
        # optimize なし・高品質でファイルを大きく保つ
        img.save(path, "JPEG", quality=quality, optimize=False, subsampling=0)
    else:
        # compress_level 低めで容量を確保（テスト用の「大きい画像」）
        img.save(path, "PNG", compress_level=1, optimize=False)
    return path


def style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN


def set_widths(ws, widths: dict):
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


def build_cover(ws, side: str):
    set_widths(ws, {"A": 4, "B": 28, "C": 55, "D": 24})
    ws["B2"] = "DiffXL 大画像ストレステスト"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "左 (基準)" if side == "left" else "右 (比較対象)"
    ws["B3"].font = SUB_FONT

    rows = [
        ("用途", "高解像度埋め込み画像の抽出・比較・MiniMap"),
        ("版", "L-1.0" if side == "left" else "R-1.1"),
        ("画像", "FHD / QHD / 4K 級 PNG・JPEG を複数枚"),
        ("期待", "比較がタイムアウトせず画像差分が検出されること"),
    ]
    for i, (k, v) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=k).font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="D9E2F3")
        ws.cell(row=r, column=2).border = THIN
        cell = ws.cell(row=r, column=3, value=v)
        cell.font = BODY_FONT
        cell.border = THIN
        if side == "right" and k == "版":
            cell.fill = DIFF_HINT_FILL

    ws["B10"] = "シート"
    ws["C10"] = "検証内容"
    style_header(ws, 10, 3)
    for c in range(2, 4):
        ws.cell(row=10, column=c).fill = HEADER_FILL
        ws.cell(row=10, column=c).font = HEADER_FONT

    sheets = [
        ("表紙", "版テキスト差分"),
        ("売上サマリ", "テキスト差分（auto-live 互換）"),
        ("製品カタログ", "大画像比較（同一/差分/片側）"),
        ("長い一覧", "MiniMap・同期スクロール"),
        ("レイアウト確認", "表示忠実性（小差分）"),
    ]
    for i, (a, b) in enumerate(sheets):
        r = 11 + i
        ws.cell(row=r, column=2, value=a).font = BODY_FONT
        ws.cell(row=r, column=3, value=b).font = BODY_FONT
        for c in range(2, 4):
            ws.cell(row=r, column=c).border = THIN
            if i % 2:
                ws.cell(row=r, column=c).fill = ALT_ROW

    note = ws["B17"]
    note.value = "※ オレンジセルは意図的差分の目印。画像本体は高解像度のまま埋め込み（表示サイズは縮小）。"
    note.fill = NOTE_FILL
    note.font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")


def build_sales(ws, side: str):
    """auto-live の probeSheets 互換: 売上サマリ。"""
    set_widths(ws, {"A": 8, "B": 14, "C": 14, "D": 10, "E": 12, "F": 12, "G": 18})
    ws["A1"] = "月次売上（テキスト差分・MiniMap 用）"
    ws["A1"].font = TITLE_FONT
    headers = ["No", "年月", "部門", "数量", "売上", "担当", "備考"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 7)

    data_l = [
        (1, "2026-01", "東日本", 100, 3000, "佐藤", "OK"),
        (2, "2026-02", "西日本", 90, 2700, "鈴木", ""),
        (3, "2026-03", "中部", 110, 3300, "田中", "計画"),
        (4, "2026-04", "東日本", 95, 2850, "佐藤", ""),
        (5, "2026-05", "西日本", 120, 3600, "鈴木", "大型"),
    ]
    data_r = [
        (1, "2026-01", "東日本", 100, 3000, "佐藤", "OK"),
        (2, "2026-02", "西日本", 98, 2950, "鈴木", "修正"),  # diff
        (3, "2026-03", "中部", 110, 3300, "田中", "計画"),
        (4, "2026-04", "東日本", 95, 3000, "高橋", ""),  # diff
        (5, "2026-05", "西日本", 125, 3750, "鈴木", "大型増"),  # diff
    ]
    data = data_l if side == "left" else data_r
    for i, row in enumerate(data):
        r = 4 + i
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = THIN
        if side == "right" and i in (1, 3, 4):
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = DIFF_HINT_FILL

    ws["A10"] = "ANCHOR_SALES"
    ws["A10"].font = MONO_FONT
    ws["B10"] = "共通アンカー" if side == "left" else "共通アンカー（右メモあり）"
    if side == "right":
        ws["C10"] = "right note"
        ws["C10"].fill = DIFF_HINT_FILL


def build_catalog(ws, side: str, images: dict):
    """大画像カタログ。表示は縮小、埋め込みファイルはフル解像度。"""
    set_widths(ws, {"A": 4, "B": 18, "C": 42, "D": 14, "E": 48})
    ws["B1"] = "製品カタログ（大画像比較）"
    ws["B1"].font = TITLE_FONT
    ws["B2"] = "埋め込み画像はフル解像度。セル上の表示サイズのみ縮小しています。"
    ws["B2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    headers = ["画像ID", "説明", "状態", "プレビュー"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 5)
    for c in range(2, 6):
        ws.cell(row=3, column=c).fill = HEADER_FILL
        ws.cell(row=3, column=c).font = HEADER_FONT

    # key_left, key_right, id, desc, st_l, st_r, display_w, display_h
    products = [
        ("same_fhd", "same_fhd", "BIG-A", "FHD 同一 PNG", "同一", "同一", 220, 124),
        ("base_qhd", "mod_qhd", "BIG-B", "QHD PNG（右は MOD スタンプ）", "基準", "内容差分", 240, 135),
        ("left_only_4k", None, "BIG-C", "4K PNG 左のみ", "左のみ", "—", 200, 150),
        (None, "right_only_4k", "BIG-D", "HD PNG 右のみ（左4Kと寸法差で誤ペア防止）", "—", "右のみ", 200, 150),
        ("same_jpg", "same_jpg", "BIG-E", "QHD JPEG 同一", "同一", "同一", 220, 124),
        ("base_wide", "mod_wide", "BIG-F", "ワイド 2560x1080（内容差分）", "基準", "内容差分", 260, 110),
        ("multi_a", "multi_a", "BIG-G", "複数画像並び A（同一）", "同一", "同一", 180, 100),
        ("multi_b", "multi_b_mod", "BIG-H", "複数画像並び B（差分）", "基準", "内容差分", 180, 100),
    ]

    row = 4
    for key_l, key_r, pid, desc, st_l, st_r, dw, dh in products:
        status = st_l if side == "left" else st_r
        key = key_l if side == "left" else key_r
        ws.row_dimensions[row].height = max(90, dh * 0.75)
        ws.cell(row=row, column=2, value=pid).font = Font(name="Consolas", size=11, bold=True)
        ws.cell(row=row, column=3, value=desc).font = BODY_FONT
        st = ws.cell(row=row, column=4, value=status)
        st.font = BODY_FONT
        st.alignment = Alignment(horizontal="center", vertical="center")
        if status in ("内容差分", "左のみ", "右のみ"):
            st.fill = WARN_FILL
        elif status == "同一":
            st.fill = OK_FILL
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = THIN
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center", wrap_text=True)

        if key and key in images:
            xl = XLImage(str(images[key]))
            xl.width = dw
            xl.height = dh
            ws.add_image(xl, f"E{row}")
        row += 1

    note_row = row + 1
    ws.cell(row=note_row, column=2).value = (
        "検証: BIG-A/E/G 同一、BIG-B/F/H 内容差分、BIG-C 左のみ、BIG-D 右のみ。"
        " 大画像でも比較が完了し MiniMap に画像差分が載ること。"
    )
    ws.cell(row=note_row, column=2).fill = NOTE_FILL
    ws.cell(row=note_row, column=2).font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    ws.cell(row=note_row + 2, column=2, value="ANCHOR_TEXT_CATALOG_SECTION").font = MONO_FONT
    ws.cell(row=note_row + 2, column=3, value="左右同一アンカー").font = BODY_FONT


def build_long_list(ws, side: str):
    set_widths(ws, {"A": 8, "B": 12, "C": 28, "D": 10, "E": 12, "F": 20})
    ws["A1"] = "長い一覧（MiniMap / 同期スクロール）"
    ws["A1"].font = TITLE_FONT
    headers = ["行", "コード", "品名", "単位", "単価", "メモ"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)

    n = 80
    for i in range(1, n + 1):
        r = 4 + i
        code = f"L-{i:04d}"
        name = f"大画像試験品目 {i:03d}"
        unit = "式" if i % 5 == 0 else "個"
        price = 1000 + (i * 41) % 8000
        memo = ""
        if side == "right":
            if i == 3:
                name = "大画像試験品目 003（改訂）"
                memo = "名称"
            elif i == 40:
                price = price + 777
                memo = "単価"
            elif i == 41:
                memo = "右のみ"
            elif i == 75:
                name = "大画像試験品目 075★"
                unit = "セット"
                memo = "下部"
        ws.cell(row=r, column=1, value=i).font = BODY_FONT
        ws.cell(row=r, column=2, value=code).font = MONO_FONT
        ws.cell(row=r, column=3, value=name).font = BODY_FONT
        ws.cell(row=r, column=4, value=unit).font = BODY_FONT
        ws.cell(row=r, column=5, value=price).font = BODY_FONT
        ws.cell(row=r, column=6, value=memo).font = BODY_FONT
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN
            if i % 2 == 0:
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")
            if side == "right" and i in (3, 40, 41, 75):
                ws.cell(row=r, column=c).fill = DIFF_HINT_FILL

    ws.cell(row=86, column=1).value = "検証: 行3・40・41・75 の差分が MiniMap に出ること。"
    ws.cell(row=86, column=1).fill = NOTE_FILL


def build_layout(ws, side: str):
    set_widths(ws, {"A": 6, "B": 24, "C": 20, "D": 16})
    ws["B1"] = "レイアウト確認"
    ws["B1"].font = TITLE_FONT
    ws["B3"] = "大画像ブックでも行高・フォントが崩れないこと"
    ws["B3"].font = BODY_FONT
    ws["B5"] = "共通テキスト"
    ws["C5"] = "layout-ok"
    if side == "right":
        ws["B7"] = "右のみレイアウトメモ"
        ws["B7"].fill = DIFF_HINT_FILL
        ws["B7"].font = BODY_FONT


def create_workbook(side: str, images: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "表紙"
    build_cover(ws, side)

    ws = wb.create_sheet("売上サマリ")
    build_sales(ws, side)

    ws = wb.create_sheet("製品カタログ")
    build_catalog(ws, side, images)

    ws = wb.create_sheet("長い一覧")
    build_long_list(ws, side)

    ws = wb.create_sheet("レイアウト確認")
    build_layout(ws, side)

    return wb


def file_info(path: Path) -> str:
    return f"{path.name}  {path.stat().st_size:,} bytes ({path.stat().st_size / 1024 / 1024:.2f} MB)"


def main():
    t0 = time.time()
    print("Generating large media...")

    # FHD same
    same_fhd = make_large_image(
        MEDIA / "same_fhd.png", (1920, 1080), "BIG-A FHD", (25, 55, 100), (70, 140, 220)
    )
    # QHD base / mod
    base_qhd = make_large_image(
        MEDIA / "base_qhd.png", (2560, 1440), "BIG-B QHD", (30, 80, 60), (80, 170, 100)
    )
    mod_qhd = make_large_image(
        MEDIA / "mod_qhd.png",
        (2560, 1440),
        "BIG-B QHD",
        (30, 80, 60),
        (80, 170, 100),
        stamp="MOD",
        noise_seed=42,
    )
    # 片側のみ: 寸法を大きくずらして誤ペア（内容比較）にならないようにする
    # 左のみ = 4K、右のみ = 小〜中解像度 → DiffEngine の面積比ヒューリスティックで ImageOnly*
    left_only_4k = make_large_image(
        MEDIA / "left_only_4k.png",
        (3840, 2160),
        "BIG-C 4K LEFT",
        (110, 40, 35),
        (200, 90, 70),
    )
    right_only_4k = make_large_image(
        MEDIA / "right_only_hd.png",
        (1280, 720),
        "BIG-D HD RIGHT",
        (80, 40, 110),
        (170, 100, 210),
    )
    # QHD JPEG same
    same_jpg = make_large_image(
        MEDIA / "same_qhd.jpg",
        (2560, 1440),
        "BIG-E JPEG",
        (50, 50, 50),
        (180, 180, 180),
        fmt="JPEG",
        quality=85,
    )
    # wide base / mod
    base_wide = make_large_image(
        MEDIA / "base_wide.png", (2560, 1080), "BIG-F WIDE", (20, 40, 80), (60, 120, 200)
    )
    mod_wide = make_large_image(
        MEDIA / "mod_wide.png",
        (2560, 1080),
        "BIG-F WIDE",
        (20, 40, 80),
        (60, 120, 200),
        stamp="CHG",
        noise_seed=99,
    )
    multi_a = make_large_image(
        MEDIA / "multi_a.png", (1600, 900), "BIG-G", (40, 40, 70), (100, 100, 180)
    )
    multi_b = make_large_image(
        MEDIA / "multi_b.png", (1600, 900), "BIG-H", (70, 40, 40), (180, 100, 90)
    )
    multi_b_mod = make_large_image(
        MEDIA / "multi_b_mod.png",
        (1600, 900),
        "BIG-H",
        (70, 40, 40),
        (180, 100, 90),
        stamp="DIFF",
        noise_seed=7,
    )

    images_left = {
        "same_fhd": same_fhd,
        "base_qhd": base_qhd,
        "left_only_4k": left_only_4k,
        "same_jpg": same_jpg,
        "base_wide": base_wide,
        "multi_a": multi_a,
        "multi_b": multi_b,
    }
    images_right = {
        "same_fhd": same_fhd,
        "mod_qhd": mod_qhd,
        "right_only_4k": right_only_4k,
        "same_jpg": same_jpg,
        "mod_wide": mod_wide,
        "multi_a": multi_a,
        "multi_b_mod": multi_b_mod,
    }

    left_path = OUT_DIR / "large_image_left.xlsx"
    right_path = OUT_DIR / "large_image_right.xlsx"

    print("Building workbooks...")
    wb_l = create_workbook("left", images_left)
    wb_l.save(left_path)
    wb_r = create_workbook("right", images_right)
    wb_r.save(right_path)

    # media sizes
    print("\nMedia files:")
    total_media = 0
    for p in sorted(MEDIA.glob("*")):
        if p.is_file():
            total_media += p.stat().st_size
            print(" ", file_info(p))
    print(f"  media total: {total_media / 1024 / 1024:.2f} MB")

    print("\nWorkbooks:")
    print(" ", file_info(left_path))
    print(" ", file_info(right_path))
    print(f"\nDone in {time.time() - t0:.1f}s")
    print("Expected image diffs: BIG-B, BIG-C(left-only), BIG-D(right-only), BIG-F, BIG-H (+ text diffs)")


if __name__ == "__main__":
    main()
