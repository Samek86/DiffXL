# -*- coding: utf-8 -*-
"""
DiffXL ストレステスト用 .xlsx を生成する。

出力（親ディレクトリ 30_参考資料/samples/）:
  - stress_suite_left.xlsx
  - stress_suite_right.xlsx

シート:
  1) 長大一覧  … 約 1000 行。表・画像・散在テキスト・片側のみ・微差
  2) 画面キャプチャ … 画面全体サイズのキャプチャ風大画像 5 枚（一部領域だけ差）

再生成:
  python 30_参考資料/samples/_gen/create_stress_samples.py
"""

from __future__ import annotations

import hashlib
import random
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT.parent
MEDIA = ROOT / "media_stress"
MEDIA.mkdir(parents=True, exist_ok=True)

# 画面キャプチャ相当（フル HD 寄り。ファイル肥大を抑えるため 1600x900）
SCREEN_SIZE = (1600, 900)
# 長大一覧に埋め込む小さめサムネ
THUMB_SIZE = (180, 100)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Yu Gothic UI", size=14, bold=True, color="1F4E79")
BODY_FONT = Font(name="Yu Gothic UI", size=10)
MONO_FONT = Font(name="Consolas", size=9)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
DIFF_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
ALT_ROW = PatternFill("solid", fgColor="F2F7FB")
LEFT_ONLY_FILL = PatternFill("solid", fgColor="FCE4D6")
RIGHT_ONLY_FILL = PatternFill("solid", fgColor="DDEBF7")


def _font(size: int):
    for name in ("segoeui.ttf", "arial.ttf", "YuGothM.ttc", "meiryo.ttc", "YuGothR.ttc"):
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def make_thumb(path: Path, label: str, bg: tuple, accent: tuple, stamp: str | None = None) -> Path:
    """一覧用の小さなラベル画像。"""
    w, h = THUMB_SIZE
    img = Image.new("RGB", (w, h), bg)
    draw = ImageDraw.Draw(img)
    draw.rectangle([3, 3, w - 4, h - 4], outline=accent, width=3)
    draw.rectangle([3, 3, w - 4, 28], fill=accent)
    f = _font(16)
    bbox = draw.textbbox((0, 0), label, font=f)
    tw = bbox[2] - bbox[0]
    draw.text(((w - tw) / 2, 6), label, fill=(255, 255, 255), font=f)
    f2 = _font(12)
    draw.text((8, 40), f"{w}x{h}", fill=accent, font=f2)
    if stamp:
        draw.rectangle([w // 2, h // 2, w - 6, h - 6], fill=(220, 50, 50))
        draw.text((w // 2 + 8, h // 2 + 10), stamp, fill=(255, 255, 255), font=f2)
    img.save(path, "PNG", optimize=True)
    return path


def make_screen_capture(
    path: Path,
    app_title: str,
    panel_color: tuple,
    seed: int,
    mods: list[dict] | None = None,
) -> Path:
    """
    デスクトップ画面キャプチャ風の大画像。
    mods: [{"kind":"rect"|"text"|"badge", ...}] で部分差分を描画。
    """
    w, h = SCREEN_SIZE
    # 背景: 薄い壁紙グラデ + タスクバー
    img = Image.new("RGB", (w, h), (32, 48, 72))
    draw = ImageDraw.Draw(img)

    # 壁紙グラデーション
    for y in range(h - 48):
        t = y / max(1, h - 48)
        r = int(40 + 30 * t)
        g = int(70 + 40 * t)
        b = int(110 + 50 * t)
        draw.line([(0, y), (w, y)], fill=(r, g, b))

    # タスクバー
    draw.rectangle([0, h - 48, w, h], fill=(20, 20, 28))
    draw.rectangle([8, h - 40, 48, h - 8], fill=(0, 120, 215))
    f_small = _font(14)
    draw.text((60, h - 34), "10:24  DiffXL Stress", fill=(220, 220, 220), font=f_small)

    # メインウィンドウ（スクリーン中央）
    mx, my, mw, mh = 80, 40, w - 160, h - 120
    draw.rectangle([mx, my, mx + mw, my + mh], fill=(245, 246, 248), outline=(60, 60, 70), width=2)
    # タイトルバー
    draw.rectangle([mx, my, mx + mw, my + 36], fill=panel_color)
    f_title = _font(20)
    draw.text((mx + 14, my + 8), app_title, fill=(255, 255, 255), font=f_title)
    # ウィンドウボタン
    for i, col in enumerate([(232, 17, 35), (255, 185, 0), (40, 180, 80)]):
        bx = mx + mw - 20 - i * 28
        draw.ellipse([bx, my + 10, bx + 14, my + 24], fill=col)

    # 左ナビ
    draw.rectangle([mx, my + 36, mx + 180, my + mh], fill=(236, 239, 244))
    f_nav = _font(15)
    for i, name in enumerate(["ダッシュボード", "レポート", "設定", "ログ", "エクスポート"]):
        yy = my + 56 + i * 36
        draw.rectangle([mx + 8, yy, mx + 168, yy + 28], fill=(panel_color if i == 0 else (220, 224, 230)))
        draw.text((mx + 18, yy + 5), name, fill=(255, 255, 255) if i == 0 else (40, 40, 50), font=f_nav)

    # コンテンツカード
    cards = [
        (mx + 200, my + 56, "売上サマリ", f"seed={seed} / KPI-A"),
        (mx + 520, my + 56, "注文件数", f"N={1000 + seed * 3}"),
        (mx + 840, my + 56, "エラー率", f"{(seed % 7) + 1}.2%"),
        (mx + 200, my + 220, "グラフ領域", "chart placeholder"),
        (mx + 680, my + 220, "直近イベント", "event stream"),
    ]
    f_card = _font(18)
    f_body = _font(14)
    for x, y, title, body in cards:
        cw, ch = 280, 140
        if title == "グラフ領域":
            cw, ch = 440, 280
        if title == "直近イベント":
            cw, ch = 360, 280
        draw.rectangle([x, y, x + cw, y + ch], fill=(255, 255, 255), outline=(200, 205, 215), width=2)
        draw.rectangle([x, y, x + cw, y + 32], fill=(panel_color[0], panel_color[1], panel_color[2]))
        draw.text((x + 10, y + 6), title, fill=(255, 255, 255), font=f_card)
        draw.text((x + 12, y + 50), body, fill=(50, 50, 60), font=f_body)
        # 簡易チャート線
        if "グラフ" in title:
            pts = []
            for i in range(12):
                px = x + 20 + i * 32
                py = y + 200 - ((i * seed * 13 + 17) % 90)
                pts.append((px, py))
            if len(pts) >= 2:
                draw.line(pts, fill=panel_color, width=3)

    # ステータスバー
    draw.rectangle([mx, my + mh - 28, mx + mw, my + mh], fill=(230, 233, 238))
    draw.text((mx + 10, my + mh - 22), f"capture {w}x{h} · id={seed}", fill=(80, 80, 90), font=f_small)

    # 決定論ノイズ（同一 seed は同一、比較で「ほぼ同じ」を保証）
    # bulk=True 相当: 圧縮しにくいノイズを足してファイルを大きくし、OpenCV 負荷も現実的にする
    try:
        import numpy as np

        rng = np.random.RandomState(seed * 9973 + 11)
        arr = np.array(img, dtype=np.float32)
        # 画面全体に中程度のノイズ（PNG が小さくなりすぎない）
        noise = rng.randint(0, 48, (h, w, 3), dtype=np.uint8).astype(np.float32)
        arr = (arr * 0.72 + noise * 0.28).clip(0, 255).astype(np.uint8)
        img = Image.fromarray(arr, "RGB")
        draw = ImageDraw.Draw(img)
        # UI 枠をもう一度はっきり描き直す（ノイズで薄くなった分）
        draw.rectangle([mx, my, mx + mw, my + mh], outline=(60, 60, 70), width=2)
        draw.rectangle([mx, my, mx + mw, my + 36], fill=panel_color)
        draw.text((mx + 14, my + 8), app_title, fill=(255, 255, 255), font=f_title)
    except Exception:
        pass

    # 部分差分
    for mod in mods or []:
        kind = mod.get("kind", "rect")
        if kind == "rect":
            box = mod["box"]  # x0,y0,x1,y1 absolute
            draw.rectangle(box, fill=mod.get("fill", (220, 40, 40)))
            if "label" in mod:
                draw.text((box[0] + 8, box[1] + 8), mod["label"], fill=(255, 255, 255), font=f_card)
        elif kind == "text":
            draw.text(mod["xy"], mod["text"], fill=mod.get("fill", (200, 30, 30)), font=f_title)
        elif kind == "badge":
            x, y = mod["xy"]
            draw.ellipse([x, y, x + 64, y + 64], fill=mod.get("fill", (255, 140, 0)))
            draw.text((x + 12, y + 18), mod.get("label", "!"), fill=(255, 255, 255), font=f_card)

    # フッタ識別
    draw.text((16, 16), path.stem, fill=(255, 255, 255), font=f_small)
    # optimize=False + 高エントロピーでサイズ確保（だいたい 1〜3MB/枚）
    img.save(path, "PNG", optimize=False, compress_level=3)
    return path


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_border_range(ws, r0, r1, c0, c1):
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).font = BODY_FONT


# 長大表の非対称プラン用 seed（左右ブックで同一結果）
LONG_TABLE_SEED = 20260813


def _pick_clusters(rng: random.Random, n_shared: int, n_clusters: int, size_lo: int, size_hi: int) -> set[int]:
    """重複しにくい連続 ID クラスタを選ぶ。"""
    chosen: set[int] = set()
    for _ in range(n_clusters):
        size = rng.randint(size_lo, size_hi)
        max_start = max(5, n_shared - size - 5)
        for _try in range(16):
            start = rng.randint(5, max_start)
            block = set(range(start, min(n_shared, start + size)))
            if len(block & chosen) <= max(1, size // 4):
                chosen |= block
                break
    return chosen


def plan_asymmetric_long_rows(n_shared: int = 1000) -> dict:
    """
    左右で行数が自然にずれる長大表プラン（決定論）。

    - 共通候補 S-0001..S-n
    - 片側だけ残る共通行（相手から削除）= クラスタ単位
    - 片側だけ追加の挿入行 = 別クラスタ（複数行まとめて）
    - 左右で削除・挿入の量を非対称にする（片方の方が明らかに少ない）
    """
    rng = random.Random(LONG_TABLE_SEED)

    # 右にだけ残る共通（左から削除）— 小さめクラスタ
    right_exclusive_shared = _pick_clusters(rng, n_shared, n_clusters=3, size_lo=2, size_hi=10)
    # 左にだけ残る共通（右から削除）— 大きめに消し、右の総行数を明確に減らす
    left_exclusive_shared = _pick_clusters(rng, n_shared, n_clusters=5, size_lo=15, size_hi=40)
    # 両方から消すと空になるので交差は右 exclusive から除去（左を残す）
    both = left_exclusive_shared & right_exclusive_shared
    right_exclusive_shared -= both

    # 挿入: アンカー shared index の直前に N 行（片側のみ・まとまったブロック）
    def place_inserts(n_blocks: int, size_lo: int, size_hi: int) -> dict[int, int]:
        """anchor_index -> count"""
        m: dict[int, int] = {}
        used: set[int] = set()
        for _ in range(n_blocks):
            for _try in range(24):
                anchor = rng.randint(20, n_shared - 20)
                if any(abs(anchor - u) < 20 for u in used):
                    continue
                used.add(anchor)
                m[anchor] = rng.randint(size_lo, size_hi)
                break
        return m

    # 左のみ追加は少なめ、右のみ追加は中程度（削除で減った右を少し補うが総数は左より少なく）
    left_inserts = place_inserts(n_blocks=2, size_lo=3, size_hi=12)
    right_inserts = place_inserts(n_blocks=3, size_lo=5, size_hi=16)

    left_specs: list[tuple] = []
    right_specs: list[tuple] = []
    left_ins_seq = 0
    right_ins_seq = 0

    for i in range(n_shared):
        # 挿入（この shared の前）
        if i in left_inserts:
            for _ in range(left_inserts[i]):
                left_ins_seq += 1
                left_specs.append(("left_only", left_ins_seq, i))
        if i in right_inserts:
            for _ in range(right_inserts[i]):
                right_ins_seq += 1
                right_specs.append(("right_only", right_ins_seq, i))

        # 共通行（片側削除あり）
        # right_exclusive_shared: 右のみ（左に無い）
        # left_exclusive_shared: 左のみ（右に無い）
        if i not in right_exclusive_shared:
            left_specs.append(("shared", i))
        if i not in left_exclusive_shared:
            right_specs.append(("shared", i))

    stats = {
        "left_count": len(left_specs),
        "right_count": len(right_specs),
        "left_inserts": left_ins_seq,
        "right_inserts": right_ins_seq,
        "left_exclusive_shared": len(left_exclusive_shared),
        "right_exclusive_shared": len(right_exclusive_shared),
        "shared_both": n_shared - len(left_exclusive_shared | right_exclusive_shared),
    }
    return {"left": left_specs, "right": right_specs, "stats": stats}


def build_long_sheet(wb: Workbook, side: str) -> None:
    """
    約 1000 行の長大シート。
    side: 'L' or 'R'
    """
    is_left = side == "L"
    ws = wb.create_sheet("長大一覧")
    ws["A1"] = f"長大ストレステスト（{'左・基準' if is_left else '右・比較'}）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "共通≈1000行候補 + 非対称な削除クラスタ + 片側挿入クラスタ（左右で行数が大きく異なり得る）。"
        " 「この側になし」・MiniMap・スクロール同期の負荷確認用。"
    )
    ws["A2"].font = BODY_FONT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # --- サマリ表（ボーダー付き） row 4-10 ---
    ws["A4"] = "サマリ表（月次）"
    ws["A4"].font = Font(name="Yu Gothic UI", size=11, bold=True, color="1F4E79")
    headers = ["No", "年月", "部門", "数量", "売上", "担当", "状態", "備考"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, 8)

    depts = ["東日本", "西日本", "中部", "九州", "北海道"]
    owners = ["佐藤", "鈴木", "田中", "高橋", "伊藤", "渡辺"]
    for i in range(5):
        r = 6 + i
        ym = f"2026-{i + 1:02d}"
        qty = 100 + i * 7
        sales = qty * 30
        owner = owners[i % len(owners)]
        status = "確定"
        note = ""
        # 左右で少しずつ差
        if not is_left:
            if i == 1:
                qty += 3
                sales = qty * 30
                note = "修正"
            if i == 3:
                owner = "山本"
            if i == 4:
                status = "確認中"
                note = "右のみ注記"
        else:
            if i == 4:
                note = "左メモ"

        vals = [i + 1, ym, depts[i % len(depts)], qty, sales, owner, status, note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = ALT_ROW
            if note and c == 8 and not is_left and i in (1, 4):
                cell.fill = DIFF_FILL
            if note and c == 8 and is_left and i == 4:
                cell.fill = NOTE_FILL

    apply_border_range(ws, 5, 10, 1, 8)

    # 片側のみブロック
    if is_left:
        ws["A12"] = "【左のみ】監査メモ: 基準ファイル固有の注意書き（右には存在しない）"
        ws["A12"].fill = LEFT_ONLY_FILL
        ws["A12"].font = BODY_FONT
        ws.merge_cells("A12:H12")
    else:
        ws["A12"] = "【右のみ】改訂コメント: 比較対象側だけの追記行"
        ws["A12"].fill = RIGHT_ONLY_FILL
        ws["A12"].font = BODY_FONT
        ws.merge_cells("A12:H12")

    ws["A13"] = "ANCHOR_LONG_TOP"
    ws["B13"] = "共通アンカー" if is_left else "共通アンカー（右メモあり）"
    ws["A13"].font = MONO_FONT
    ws["B13"].font = BODY_FONT
    if not is_left:
        ws["C13"] = "right-top-note"
        ws["C13"].fill = DIFF_FILL

    # --- メイン長表 header row 15 ---
    # 共通行 ≈1000 + 非対称な削除クラスタ + 片側挿入クラスタ（件数・位置は seed 固定の疑似乱数）
    start = 15
    long_headers = ["行", "コード", "品名", "カテゴリ", "数量", "単価", "金額", "担当", "更新日", "フラグ", "メモ"]
    for i, h in enumerate(long_headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(long_headers))

    n_shared = 1000
    cats = ["部品", "消耗品", "機器", "サービス", "ライセンス"]
    flags = ["", "注意", "優先", "保留"]

    # 左右で同じプランを使う（決定論）
    plan = plan_asymmetric_long_rows(n_shared)
    side_specs = plan["left"] if is_left else plan["right"]

    def build_shared_vals(i: int, for_left: bool) -> list:
        code = f"S-{i + 1:04d}"
        name = f"試験品目 {i + 1:04d}"
        cat = cats[i % len(cats)]
        qty = 1 + (i % 20)
        price = 100 + (i * 17) % 900
        owner = owners[i % len(owners)]
        day = 1 + (i % 28)
        flag = flags[i % len(flags)]
        memo = ""

        differ = (i % 20 == 7) or (i % 37 == 0) or (i in (10, 100, 250, 500, 750, 999))
        if for_left:
            if differ:
                if i % 20 == 7:
                    qty += 1
                    memo = "数量微差"
                elif i % 37 == 0:
                    name = f"試験品目 {i + 1:04d}（改訂候補）"
                    flag = "注意"
                else:
                    price += 5
                    memo = "単価調整"
        else:
            if differ:
                if i % 20 == 7:
                    qty += 2
                    memo = "数量改訂"
                elif i % 37 == 0:
                    name = f"試験品目 {i + 1:04d}（確定名）"
                    flag = "優先"
                else:
                    price += 8
                    memo = "単価確定"

        if i % 51 == 3:
            flag = ""

        if not for_left and i == 120:
            name = ""
            memo = "品名クリア（右）"
        if for_left and i == 121:
            memo = "左だけ長いコメント: " + ("あ" * 40)

        amount = qty * price
        return [
            i + 1,
            code,
            name,
            cat,
            qty,
            price,
            amount,
            owner,
            f"2026-06-{day:02d}",
            flag,
            memo,
        ]

    def build_onesided_vals(side_tag: str, seq: int, anchor: int) -> list:
        """片側のみ行（L-/R- コードで相手とマッチしない）。"""
        code = f"{side_tag}-INS-{anchor + 1:04d}-{seq:03d}"
        name = f"{'左のみ' if side_tag == 'L' else '右のみ'}追加 {anchor + 1:04d}-{seq:03d}"
        cat = "片側のみ"
        qty = 10 + (seq % 9)
        price = 200 + (seq * 13) % 500
        amount = qty * price
        owner = owners[seq % len(owners)]
        day = 1 + (seq % 28)
        flag = "ONLY"
        memo = f"{'左' if side_tag == 'L' else '右'}のみクラスタ（この側になし検証）"
        return [0, code, name, cat, qty, price, amount, owner, f"2026-06-{day:02d}", flag, memo]

    side_rows: list[list] = []
    for spec in side_specs:
        kind = spec[0]
        if kind == "shared":
            side_rows.append(build_shared_vals(spec[1], for_left=is_left))
        elif kind == "left_only":
            side_rows.append(build_onesided_vals("L", spec[1], spec[2]))
        elif kind == "right_only":
            side_rows.append(build_onesided_vals("R", spec[1], spec[2]))

    # 書き込み（表示用「行」列は側内連番）
    for idx, vals in enumerate(side_rows):
        r = start + 1 + idx
        row_vals = list(vals)
        row_vals[0] = idx + 1
        code = str(row_vals[1])
        onesided = code.startswith("L-") or code.startswith("R-")
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = THIN
            if idx % 2 == 1 and not onesided:
                cell.fill = ALT_ROW
            if onesided:
                cell.fill = LEFT_ONLY_FILL if is_left else RIGHT_ONLY_FILL
            elif row_vals[10] and c == 11:
                cell.fill = DIFF_FILL
                if "長いコメント" in str(row_vals[10]):
                    cell.fill = LEFT_ONLY_FILL

    end_data = start + len(side_rows)
    ar = end_data + 2
    ws.cell(row=ar, column=1, value="ANCHOR_LONG_BOTTOM")
    ws.cell(row=ar, column=1).font = MONO_FONT
    ws.cell(row=ar, column=2, value="共通フッター" if is_left else "共通フッター（右追記）")
    if not is_left:
        ws.cell(row=ar, column=3, value="tail-right-only")
        ws.cell(row=ar, column=3).fill = RIGHT_ONLY_FILL

    note_r = ar + 2
    stats = plan["stats"]
    ws.cell(
        row=note_r,
        column=1,
        value=(
            f"【構成】共通候補 {n_shared} · "
            f"左表示 {stats['left_count']} 行（左のみ挿入 {stats['left_inserts']} · "
            f"右に無い共通 {stats['left_exclusive_shared']}）· "
            f"右表示 {stats['right_count']} 行（右のみ挿入 {stats['right_inserts']} · "
            f"左に無い共通 {stats['right_exclusive_shared']}）· "
            f"この側={len(side_rows)}"
        ),
    )
    ws.cell(row=note_r, column=1).font = BODY_FONT
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=11)

    # 列幅
    widths = [8, 12, 28, 12, 8, 10, 12, 10, 12, 8, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- サムネ画像を数箇所に配置（左右で一部差）---
    media_side = MEDIA / ("long_" + side.lower())
    media_side.mkdir(parents=True, exist_ok=True)
    palette = [
        ((30, 80, 140), (80, 160, 220)),
        ((20, 100, 60), (80, 180, 100)),
        ((120, 50, 20), (220, 120, 60)),
        ((80, 40, 120), (160, 100, 200)),
        ((40, 40, 40), (140, 140, 140)),
    ]
    # Excel 行はヘッダ=15 のため、データ行オフセットで散在
    placements = [
        (start + 20, "TH-A", 0, False),
        (start + 120, "TH-B", 1, is_left),
        (start + 300, "TH-C", 2, False),
        (start + 600, "TH-D", 3, not is_left),
        (start + 900, "TH-E", 4, False),
    ]
    if not is_left:
        placements.append((start + 450, "TH-RONLY", 0, True))
    if is_left:
        placements.append((start + 700, "TH-LONLY", 1, True))

    for row_anchor, label, pal_i, stamp in placements:
        bg, acc = palette[pal_i % len(palette)]
        stamp_txt = "MOD" if stamp else None
        if label == "TH-B" and not is_left:
            stamp_txt = "MOD"
        if label == "TH-D" and not is_left:
            stamp_txt = "NEW"
        if label in ("TH-RONLY", "TH-LONLY"):
            stamp_txt = "ONLY"

        path = media_side / f"{label}_{side}.png"
        make_thumb(path, label, bg, acc, stamp=stamp_txt)
        xl = XLImage(str(path))
        xl.width = THUMB_SIZE[0]
        xl.height = THUMB_SIZE[1]
        ws.add_image(xl, f"L{row_anchor}")
        ws.cell(row=row_anchor, column=12, value=f"[img:{label}]")
        ws.cell(row=row_anchor, column=12).font = MONO_FONT

    ws.column_dimensions["L"].width = 14
    ws.freeze_panes = "A16"


def build_screen_sheet(wb: Workbook, side: str) -> None:
    """画面キャプチャ風 大画像 5 枚。"""
    is_left = side == "L"
    ws = wb.create_sheet("画面キャプチャ")
    ws["A1"] = f"画面キャプチャ比較（{'左・基準' if is_left else '右・一部改変'}）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"フルスクリーン相当 {SCREEN_SIZE[0]}x{SCREEN_SIZE[1]} のキャプチャ風画像を 5 枚配置。"
        " 右は各キャプチャの一部領域だけが異なる（OpenCV 部分差ハイライト確認用）。"
    )
    ws["A2"].font = BODY_FONT
    ws.merge_cells("A2:F2")

    # キャプチャ定義: (id, title, panel_color, seed, right_mods)
    captures = [
        (
            "SCR-1",
            "DiffXL Dashboard",
            (31, 78, 121),
            1,
            [
                {"kind": "rect", "box": (920, 120, 1180, 250), "fill": (220, 40, 40), "label": "KPI MOD"},
            ],
        ),
        (
            "SCR-2",
            "Sales Report",
            (0, 120, 100),
            2,
            [
                {"kind": "badge", "xy": (1100, 300), "fill": (255, 140, 0), "label": "!!"},
                {"kind": "text", "xy": (300, 500), "text": "REVISED CHART", "fill": (200, 30, 30)},
            ],
        ),
        (
            "SCR-3",
            "Inventory View",
            (100, 60, 140),
            3,
            [
                # 左と同一（差分なし）→ mods 空
            ],
        ),
        (
            "SCR-4",
            "Settings Panel",
            (140, 70, 30),
            4,
            [
                {"kind": "rect", "box": (250, 400, 600, 520), "fill": (40, 40, 40), "label": "hidden"},
            ],
        ),
        (
            "SCR-5",
            "Audit Log",
            (50, 50, 60),
            5,
            [
                {"kind": "rect", "box": (700, 200, 1200, 320), "fill": (220, 40, 40), "label": "ALERT"},
                {"kind": "text", "xy": (100, 80), "text": "RIGHT ONLY BANNER", "fill": (255, 220, 0)},
            ],
        ),
    ]

    media_side = MEDIA / ("screen_" + side.lower())
    media_side.mkdir(parents=True, exist_ok=True)

    # 一覧表
    ws["A4"] = "ID"
    ws["B4"] = "タイトル"
    ws["C4"] = "期待"
    ws["D4"] = "備考"
    style_header_row(ws, 4, 4)

    expect_notes = [
        ("SCR-1", "部分差（KPI カード）", "右下カードに赤 MOD"),
        ("SCR-2", "部分差（バッジ＋文言）", "チャート付近"),
        ("SCR-3", "同一（差分なし）", "見た目一致のはず"),
        ("SCR-4", "部分差（領域マスク）", "設定パネル一部を黒塗り"),
        ("SCR-5", "部分差（アラート帯）", "複数領域"),
    ]
    for i, (cid, exp, note) in enumerate(expect_notes):
        ws.cell(row=5 + i, column=1, value=cid).border = THIN
        ws.cell(row=5 + i, column=2, value=captures[i][1]).border = THIN
        ws.cell(row=5 + i, column=3, value=exp).border = THIN
        ws.cell(row=5 + i, column=4, value=note).border = THIN
        for c in range(1, 5):
            ws.cell(row=5 + i, column=c).font = BODY_FONT

    ws["A11"] = "ANCHOR_SCREEN"
    ws["A11"].font = MONO_FONT
    ws["B11"] = "キャプチャ本体は下行から配置（表示サイズ縮小・実体は高解像度）"
    ws["B11"].font = BODY_FONT

    # 画像を縦に並べる（各ブロック約 30 行ぶん確保）
    row_cursor = 13
    for idx, (cid, title, color, seed, right_mods) in enumerate(captures):
        ws.cell(row=row_cursor, column=1, value=f"{cid} · {title}")
        ws.cell(row=row_cursor, column=1).font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=4)

        mods = None if is_left else (right_mods or None)
        # SCR-3 は左右同一
        if cid == "SCR-3":
            mods = None

        path = media_side / f"{cid}_{side}.png"
        make_screen_capture(path, f"{title} [{cid}]", color, seed=seed, mods=mods)

        xl = XLImage(str(path))
        # シート上の表示は縮小（埋め込み実体はフル解像度）
        display_w = 640
        display_h = int(display_w * SCREEN_SIZE[1] / SCREEN_SIZE[0])
        xl.width = display_w
        xl.height = display_h
        anchor = f"A{row_cursor + 1}"
        ws.add_image(xl, anchor)
        ws.cell(row=row_cursor + 1, column=6, value=f"{path.name} {SCREEN_SIZE[0]}x{SCREEN_SIZE[1]}")
        ws.cell(row=row_cursor + 1, column=6).font = MONO_FONT

        # 行の高さ確保（画像表示ぶん）
        for rr in range(row_cursor + 1, row_cursor + 28):
            ws.row_dimensions[rr].height = 18

        row_cursor += 30

    # 右のみの 6 枚目キャプチャ（片側のみ画像）
    if not is_left:
        ws.cell(row=row_cursor, column=1, value="SCR-RONLY · Right Only Screen")
        ws.cell(row=row_cursor, column=1).font = Font(name="Yu Gothic UI", size=11, bold=True, color="C00000")
        path = media_side / "SCR-RONLY_R.png"
        make_screen_capture(
            path,
            "Right Only Screen [SCR-RONLY]",
            (150, 30, 30),
            seed=99,
            mods=[{"kind": "text", "xy": (400, 400), "text": "RIGHT ONLY CAPTURE", "fill": (255, 255, 0)}],
        )
        xl = XLImage(str(path))
        xl.width = 480
        xl.height = int(480 * SCREEN_SIZE[1] / SCREEN_SIZE[0])
        ws.add_image(xl, f"A{row_cursor + 1}")

    # 左のみ
    if is_left:
        ws.cell(row=row_cursor, column=1, value="SCR-LONLY · Left Only Screen")
        ws.cell(row=row_cursor, column=1).font = Font(name="Yu Gothic UI", size=11, bold=True, color="C00000")
        path = media_side / "SCR-LONLY_L.png"
        make_screen_capture(
            path,
            "Left Only Screen [SCR-LONLY]",
            (30, 30, 120),
            seed=98,
            mods=[{"kind": "text", "xy": (400, 400), "text": "LEFT ONLY CAPTURE", "fill": (255, 255, 0)}],
        )
        xl = XLImage(str(path))
        xl.width = 480
        xl.height = int(480 * SCREEN_SIZE[1] / SCREEN_SIZE[0])
        ws.add_image(xl, f"A{row_cursor + 1}")

    for col, width in [("A", 18), ("B", 28), ("C", 28), ("D", 36), ("E", 12), ("F", 40)]:
        ws.column_dimensions[col].width = width


def build_cover(wb: Workbook, side: str) -> None:
    ws = wb.create_sheet("表紙", 0)
    is_left = side == "L"
    ws["B2"] = "DiffXL ストレステストスイート"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "左 (基準)" if is_left else "右 (比較対象)"
    ws["B3"].font = Font(name="Yu Gothic UI", size=12, bold=True, color="C00000" if not is_left else "1F4E79")
    ws["B5"] = "版"
    ws["C5"] = "L-stress-1.0" if is_left else "R-stress-1.1"
    ws["B6"] = "用途"
    ws["C6"] = "長大シート + 画面キャプチャ大画像の比較検証"
    ws["B7"] = "期待"
    ws["C7"] = "比較完了・部分差ハイライト・片側のみ検出・ストリーム対応が破綻しないこと"
    for r in range(5, 8):
        for c in range(2, 4):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).font = BODY_FONT

    ws["B10"] = "シート"
    ws["C10"] = "検証内容"
    for c in (2, 3):
        ws.cell(row=10, column=c).fill = HEADER_FILL
        ws.cell(row=10, column=c).font = HEADER_FONT
        ws.cell(row=10, column=c).border = THIN
        ws.cell(row=10, column=c).alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        ("表紙", "版テキスト差分"),
        ("長大一覧", "共通≈1000候補・非対称削除/挿入クラスタ・行数差・微差"),
        ("画面キャプチャ", "1600x900 キャプチャ風 5 枚 + 片側のみ 1 枚"),
    ]
    for i, (a, b) in enumerate(rows):
        ws.cell(row=11 + i, column=2, value=a).border = THIN
        ws.cell(row=11 + i, column=3, value=b).border = THIN
        ws.cell(row=11 + i, column=2).font = BODY_FONT
        ws.cell(row=11 + i, column=3).font = BODY_FONT

    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 56


def build_workbook(side: str) -> Workbook:
    wb = Workbook()
    # デフォルトシートを表紙に差し替え
    default = wb.active
    wb.remove(default)
    build_cover(wb, side)
    build_long_sheet(wb, side)
    build_screen_sheet(wb, side)
    return wb


def main():
    t0 = time.time()
    print("Generating stress suite samples...")
    print(f"  media: {MEDIA}")

    left_path = OUT_DIR / "stress_suite_left.xlsx"
    right_path = OUT_DIR / "stress_suite_right.xlsx"

    wb_l = build_workbook("L")
    wb_l.save(left_path)
    print(f"  wrote {left_path} ({left_path.stat().st_size / 1024 / 1024:.2f} MB)")

    wb_r = build_workbook("R")
    wb_r.save(right_path)
    print(f"  wrote {right_path} ({right_path.stat().st_size / 1024 / 1024:.2f} MB)")

    # 簡易マニフェスト
    man = OUT_DIR / "stress_suite_README.txt"
    man.write_text(
        """DiffXL stress suite samples
===========================
Files:
  stress_suite_left.xlsx
  stress_suite_right.xlsx

Sheets:
  表紙           version text diff
  長大一覧        ~1000 shared candidates with asymmetric clusters:
                   - multi-row delete clusters (side totals differ a lot)
                   - multi-row insert-only blocks (L-INS-* / R-INS-*)
                   + summary table + thumbs + cell-level diffs
  画面キャプチャ   5x 1600x900 screen-like captures (partial mods on right)
                   + SCR-LONLY (left only) / SCR-RONLY (right only)

Long-table plan seed: 20260813 (reproducible pseudo-random)

Regenerate:
  python 30_参考資料/samples/_gen/create_stress_samples.py
""",
        encoding="utf-8",
    )
    print(f"  wrote {man}")
    print(f"Done in {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
