# -*- coding: utf-8 -*-
"""
内容同期スクロール（完璧対応）用の専用サンプル .xlsx を生成する。

出力先: 親ディレクトリ (30_参考資料/samples/)
  - content_scroll_left.xlsx
  - content_scroll_right.xlsx
  - content_scroll_expected.json

中間メディア: _gen/media_content_scroll/

シート設計（固定）:

| シート名       | 目的 |
|----------------|------|
| SC_画像ギャップ | 左 2 枚・右 3 枚。右の真ん中だけ別物 → 縦ギャップ同期 |
| SC_テキスト挿入 | 右にだけ 2 行挿入。S01..S05 の ID で再連結 |
| SC_大画像span   | 画像が複数行にまたがる（行高を高くし twoCell で to を遠く） |
| SC_横同期       | 縦は同一内容、列だけ広く → 横 1:1 の確認用 |
| SC_同順異内容   | 順番は同じだが 2 枚目が大きく異なる → 誤ペア禁止 |

SC_画像ギャップ レイアウト（必須）:

  Left rows:
    3:  TEXT "SECTION_A"
    5:  IMAGE same_A   (hash shared)     span rows 5-6
    8:  IMAGE same_B   (hash shared)     span rows 8-9

  Right rows:
    3:  TEXT "SECTION_A"
    5:  IMAGE same_A                     span rows 5-6
    8:  IMAGE only_right_X  (unique)     span rows 8-10   ← 右のみ
    12: IMAGE same_B                     span rows 12-13

  期待スクロール:
    L5 ↔ R5   (same_A)
    R8-10 のあいだ → 左は L6 付近でホールド（same_A 終端）
    L8 ↔ R12  (same_B で再同期)

画像は openpyxl TwoCellAnchor（editAs=twoCell）で埋め込む。
行高は 60〜90 にして見た目でも span が分かるようにする。

再生成:
  python 30_参考資料/samples/_gen/create_content_scroll_samples.py
"""

from __future__ import annotations

import json
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT.parent
MEDIA = ROOT / "media_content_scroll"
MEDIA.mkdir(parents=True, exist_ok=True)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Yu Gothic UI", size=14, bold=True, color="1F4E79")
SUB_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="2E75B6")
BODY_FONT = Font(name="Yu Gothic UI", size=11)
MONO_FONT = Font(name="Consolas", size=10)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
DIFF_HINT_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
SECTION_FILL = PatternFill("solid", fgColor="D6EAF8")
INSERT_FILL = PatternFill("solid", fgColor="F8CBAD")
ID_FILL = PatternFill("solid", fgColor="E2EFDA")

# 行高: span が見えるようにやや高く
ROW_H_TEXT = 22
ROW_H_IMAGE = 72
ROW_H_IMAGE_TALL = 90


def _font(size: int):
    for name in ("segoeui.ttf", "arial.ttf", "YuGothM.ttc", "meiryo.ttc"):
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def make_badge(
    path: Path,
    size: tuple[int, int],
    label: str,
    bg: tuple[int, int, int],
    accent: tuple[int, int, int],
    stamp: str | None = None,
) -> Path:
    """テスト用 PNG を生成する。"""
    w, h = size
    img = Image.new("RGB", (w, h), bg)
    draw = ImageDraw.Draw(img)
    draw.rectangle([4, 4, w - 5, h - 5], outline=accent, width=3)
    draw.rectangle([4, 4, w - 5, max(28, h // 4)], fill=accent)
    font = _font(max(18, min(36, h // 3)))
    bbox = draw.textbbox((0, 0), label, font=font)
    tw = bbox[2] - bbox[0]
    draw.text(((w - tw) / 2, 6), label, fill=(255, 255, 255), font=font)
    meta = f"{w}x{h}"
    f2 = _font(14)
    draw.text((10, h - 24), meta, fill=accent, font=f2)
    if stamp:
        sw, sh = max(60, w // 3), max(28, h // 4)
        x0, y0 = w - sw - 10, h - sh - 10
        draw.rectangle([x0, y0, x0 + sw, y0 + sh], fill=(220, 40, 40))
        fs = _font(max(16, sh // 2))
        bb = draw.textbbox((0, 0), stamp, font=fs)
        tw3, th3 = bb[2] - bb[0], bb[3] - bb[1]
        draw.text(
            (x0 + (sw - tw3) / 2, y0 + (sh - th3) / 2),
            stamp,
            fill=(255, 255, 255),
            font=fs,
        )
        draw.polygon([(8, h // 2), (48, h // 2), (8, h // 2 + 40)], fill=(255, 210, 0))
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path, "PNG")
    return path


def set_widths(ws, widths: dict):
    for k, v in widths.items():
        ws.column_dimensions[k].width = v


def style_header_row(ws, row: int, cols: int, start_col: int = 1):
    for c in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN


def set_row_heights(ws, start: int, end: int, height: float):
    for r in range(start, end + 1):
        ws.row_dimensions[r].height = height


def add_image_twocell(
    ws,
    path: Path,
    *,
    row_start: int,
    row_end: int,
    col_start: int = 2,
    col_end: int | None = None,
    display_size: tuple[int, int] | None = None,
):
    """
    TwoCellAnchor で画像を埋め込む。
    row_start/row_end は 1-based 含む範囲（Excel 行番号）。
    OOXML marker は 0-based。to.row は row_end の次行先頭（= row_end の下端）。
    """
    if col_end is None:
        col_end = col_start + 2  # B..D 相当の幅

    xl = XLImage(str(path))
    if display_size:
        xl.width, xl.height = display_size

    # 0-based markers: from = top-left of row_start, to = top of (row_end+1)
    _from = AnchorMarker(col=col_start - 1, colOff=0, row=row_start - 1, rowOff=0)
    to = AnchorMarker(col=col_end, colOff=0, row=row_end, rowOff=0)
    xl.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
    ws.add_image(xl)
    return xl


def generate_media() -> dict[str, Path]:
    """
    画像ピクセル仕様（誤ペア防止）:
      same_A       320×120  青 + A
      same_B       320×120  緑 + B
      only_right_X 200×200  赤 + X
      only_left_Y  180×90   黄 + Y
      decoy        400×80   紫
      mod_b        320×120  same_B の改訂スタンプ
    """
    media = {
        "same_a": make_badge(
            MEDIA / "same_a.png", (320, 120), "same_A", (30, 80, 160), (70, 140, 220)
        ),
        "same_b": make_badge(
            MEDIA / "same_b.png", (320, 120), "same_B", (30, 110, 60), (70, 180, 100)
        ),
        "mod_b": make_badge(
            MEDIA / "mod_b.png",
            (320, 120),
            "same_B",
            (30, 110, 60),
            (70, 180, 100),
            stamp="MOD",
        ),
        "only_right": make_badge(
            MEDIA / "only_right.png", (200, 200), "only_X", (160, 40, 40), (220, 80, 70)
        ),
        "only_left": make_badge(
            MEDIA / "only_left.png", (180, 90), "only_Y", (180, 150, 30), (220, 190, 60)
        ),
        "decoy": make_badge(
            MEDIA / "decoy.png", (400, 80), "DECOY", (90, 40, 130), (170, 100, 210)
        ),
        # 大画像 span 用（やや大きいが扱いやすいサイズ）
        "span_big": make_badge(
            MEDIA / "span_big.png",
            (480, 280),
            "SPAN_BIG",
            (40, 50, 90),
            (90, 130, 200),
        ),
    }
    return media


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_sc_image_gap(ws, side: str, media: dict[str, Path]):
    """SC_画像ギャップ: 左 2・右 3、中央右のみ。"""
    set_widths(ws, {"A": 6, "B": 16, "C": 18, "D": 18, "E": 28})
    ws["A1"] = "SC_画像ギャップ"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "左2枚・右3枚。右中央 only_right_X 区間は左ホールド、same_B で再同期。"
        if side == "left"
        else "右中央に only_right_X（行8-10）。same_B は行12-13。"
    )
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    # row 3: SECTION_A
    ws.row_dimensions[3].height = ROW_H_TEXT
    cell = ws.cell(row=3, column=2, value="SECTION_A")
    cell.font = SUB_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN
    ws.cell(row=3, column=3, value="共通セクション見出し").font = BODY_FONT

    if side == "left":
        # same_A rows 5-6
        set_row_heights(ws, 5, 6, ROW_H_IMAGE)
        ws.cell(row=5, column=2, value="IMG same_A").font = MONO_FONT
        ws.cell(row=5, column=2).fill = OK_FILL
        add_image_twocell(ws, media["same_a"], row_start=5, row_end=6, col_start=3, col_end=5)

        # same_B rows 8-9
        set_row_heights(ws, 8, 9, ROW_H_IMAGE)
        ws.cell(row=8, column=2, value="IMG same_B").font = MONO_FONT
        ws.cell(row=8, column=2).fill = OK_FILL
        add_image_twocell(ws, media["same_b"], row_start=8, row_end=9, col_start=3, col_end=5)

        ws.cell(row=11, column=2, value="※ 左は same_A / same_B の 2 枚のみ").fill = NOTE_FILL
        ws.cell(row=11, column=2).font = Font(name="Yu Gothic UI", size=9, italic=True)
    else:
        # same_A rows 5-6
        set_row_heights(ws, 5, 6, ROW_H_IMAGE)
        ws.cell(row=5, column=2, value="IMG same_A").font = MONO_FONT
        ws.cell(row=5, column=2).fill = OK_FILL
        add_image_twocell(ws, media["same_a"], row_start=5, row_end=6, col_start=3, col_end=5)

        # only_right_X rows 8-10
        set_row_heights(ws, 8, 10, ROW_H_IMAGE)
        ws.cell(row=8, column=2, value="IMG only_right_X").font = MONO_FONT
        ws.cell(row=8, column=2).fill = WARN_FILL
        add_image_twocell(
            ws, media["only_right"], row_start=8, row_end=10, col_start=3, col_end=5
        )

        # same_B rows 12-13
        set_row_heights(ws, 12, 13, ROW_H_IMAGE)
        ws.cell(row=12, column=2, value="IMG same_B").font = MONO_FONT
        ws.cell(row=12, column=2).fill = OK_FILL
        add_image_twocell(ws, media["same_b"], row_start=12, row_end=13, col_start=3, col_end=5)

        ws.cell(row=15, column=2, value="※ 行8-10 は右のみ。同期時は左ホールド期待。").fill = NOTE_FILL
        ws.cell(row=15, column=2).font = Font(name="Yu Gothic UI", size=9, italic=True)


def build_sc_text_insert(ws, side: str):
    """
    SC_テキスト挿入: 右にだけ 2 行挿入。S01..S05 で再連結。

    Left:
      5:S01  6:data  7:S02  8:data  9:data  10:S03  11:data  12:S04  13:data  14:S05
    Right (2 insert rows after S02 / before former row 8):
      5:S01  6:data  7:S02  8:INSERT1  9:INSERT2  10:(ex8)  11:(ex9)  12:S03 ...
    """
    set_widths(ws, {"A": 6, "B": 12, "C": 28, "D": 16, "E": 24})
    ws["A1"] = "SC_テキスト挿入"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "左基準。S01..S05 がランドマーク。"
        if side == "left"
        else "S02 の直後に 2 行挿入（INSERT）。S03 は左10 ↔ 右12。"
    )
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    headers = ["行", "ID", "内容", "種別", "メモ"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 5)

    # Shared body for left (no inserts)
    left_rows = [
        (5, "S01", "セクション開始", "id", "共通"),
        (6, "", "S01 配下データ A", "data", ""),
        (7, "S02", "中間ブロック", "id", "共通"),
        (8, "", "S02 配下データ B1", "data", ""),
        (9, "", "S02 配下データ B2", "data", ""),
        (10, "S03", "再連結ポイント", "id", "L10↔R12"),
        (11, "", "S03 配下データ C", "data", ""),
        (12, "S04", "後半ブロック", "id", "共通"),
        (13, "", "S04 配下データ D", "data", ""),
        (14, "S05", "終端", "id", "共通"),
        (15, "", "S05 配下データ E", "data", ""),
    ]

    if side == "left":
        rows = left_rows
    else:
        # Insert 2 rows after row 7 (S02): right rows 8-9 are inserts
        rows = [
            (5, "S01", "セクション開始", "id", "共通"),
            (6, "", "S01 配下データ A", "data", ""),
            (7, "S02", "中間ブロック", "id", "共通"),
            (8, "INS1", "【挿入】右のみ行 1", "insert", "左に対応なし"),
            (9, "INS2", "【挿入】右のみ行 2", "insert", "左に対応なし"),
            (10, "", "S02 配下データ B1", "data", "左8相当"),
            (11, "", "S02 配下データ B2", "data", "左9相当"),
            (12, "S03", "再連結ポイント", "id", "L10↔R12"),
            (13, "", "S03 配下データ C", "data", ""),
            (14, "S04", "後半ブロック", "id", "共通"),
            (15, "", "S04 配下データ D", "data", ""),
            (16, "S05", "終端", "id", "共通"),
            (17, "", "S05 配下データ E", "data", ""),
        ]

    for r, sid, content, kind, memo in rows:
        ws.row_dimensions[r].height = ROW_H_TEXT
        ws.cell(row=r, column=1, value=r).font = MONO_FONT
        id_cell = ws.cell(row=r, column=2, value=sid if sid else None)
        id_cell.font = Font(name="Consolas", size=11, bold=True)
        ws.cell(row=r, column=3, value=content).font = BODY_FONT
        ws.cell(row=r, column=4, value=kind).font = BODY_FONT
        ws.cell(row=r, column=5, value=memo).font = BODY_FONT
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).alignment = Alignment(vertical="center")
        if kind == "id":
            id_cell.fill = ID_FILL
        elif kind == "insert":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = INSERT_FILL

    note_r = 17 if side == "left" else 19
    ws.cell(
        row=note_r,
        column=2,
        value=(
            "検証: L10(S03)→R12 / R8 挿入区間は左ホールド(≤7)。"
            if side == "left"
            else "検証: 挿入行(R8-9)スクロール中は左が S02 付近でホールド。"
        ),
    ).fill = NOTE_FILL
    ws.cell(row=note_r, column=2).font = Font(
        name="Yu Gothic UI", size=9, italic=True, color="666666"
    )


def build_sc_large_span(ws, side: str, media: dict[str, Path]):
    """SC_大画像span: twoCell で複数行にまたがる画像。"""
    set_widths(ws, {"A": 6, "B": 16, "C": 20, "D": 20, "E": 20, "F": 16})
    ws["A1"] = "SC_大画像span"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "行高を高くし twoCell で from→to を遠くする。スパン中は同一ペア内マッピング。"
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    ws.cell(row=3, column=2, value="SPAN_SECTION").font = SUB_FONT
    ws.cell(row=3, column=2).fill = SECTION_FILL

    # Large span image rows 5-10 (6 rows)
    set_row_heights(ws, 5, 10, ROW_H_IMAGE_TALL)
    ws.cell(row=5, column=2, value="IMG SPAN_BIG").font = MONO_FONT
    ws.cell(row=5, column=2).fill = OK_FILL
    add_image_twocell(
        ws,
        media["span_big"],
        row_start=5,
        row_end=10,
        col_start=3,
        col_end=6,
    )

    # Small trailing pair image — left/right same position
    set_row_heights(ws, 12, 13, ROW_H_IMAGE)
    ws.cell(row=12, column=2, value="IMG same_A (tail)").font = MONO_FONT
    add_image_twocell(ws, media["same_a"], row_start=12, row_end=13, col_start=3, col_end=5)

    # Optional minor text diff on right for visibility
    if side == "right":
        ws.cell(row=15, column=2, value="右メモ: span 下端からの再同期確認用").fill = DIFF_HINT_FILL
        ws.cell(row=15, column=2).font = BODY_FONT
    else:
        ws.cell(row=15, column=2, value="左: span 5-10 + tail 12-13").font = BODY_FONT

    ws.cell(row=17, column=2, value="検証: 画像スパン中は同一ペア内。下端到達で tail へ。").fill = NOTE_FILL
    ws.cell(row=17, column=2).font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")


def build_sc_horizontal(ws, side: str):
    """SC_横同期: 縦は同一、列だけ広く → 横 1:1。"""
    # Wide columns
    widths = {chr(ord("A") + i): 18 for i in range(12)}
    widths["A"] = 8
    set_widths(ws, widths)

    ws["A1"] = "SC_横同期"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "縦内容は左右同一。列を広げ横スクロール 1:1 を確認する。"
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    headers = [f"COL{i:02d}" for i in range(1, 12)]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 11)

    # Same vertical content both sides
    for r in range(5, 25):
        ws.row_dimensions[r].height = ROW_H_TEXT
        for c in range(1, 12):
            val = f"R{r:02d}C{c:02d}"
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = MONO_FONT
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # Marker rows for vertical identity checks
    ws.cell(row=5, column=1, value="H_TOP").font = Font(name="Consolas", size=10, bold=True)
    ws.cell(row=5, column=1).fill = ID_FILL
    ws.cell(row=14, column=1, value="H_MID").font = Font(name="Consolas", size=10, bold=True)
    ws.cell(row=14, column=1).fill = ID_FILL
    ws.cell(row=24, column=1, value="H_BOT").font = Font(name="Consolas", size=10, bold=True)
    ws.cell(row=24, column=1).fill = ID_FILL

    # Side tag only in a non-sync critical cell (note area)
    note = ws.cell(
        row=26,
        column=1,
        value=f"横同期用（{'左' if side == 'left' else '右'}）。列スクロール時に見出しが一致すること。",
    )
    note.fill = NOTE_FILL
    note.font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")


def build_sc_same_order_diff(ws, side: str, media: dict[str, Path]):
    """
    SC_同順異内容: 順番は同じだが 2 枚目が大きく異なる → 誤ペア禁止。

    1枚目: same_A（左右同一）
    2枚目: 左 same_B / 右 mod_b（内容差分・同寸法）または decoy を近くに置かない
    左のみ only_left を末尾に（任意の leftOnly）
    """
    set_widths(ws, {"A": 6, "B": 18, "C": 18, "D": 18, "E": 24})
    ws["A1"] = "SC_同順異内容"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "1枚目同一・2枚目は内容が大きく異なる。寸法差 decoy とは誤ペアしないこと。"
        if side == "left"
        else "2枚目は mod_b（MOD スタンプ）。順番は左と同じ。"
    )
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    ws.cell(row=3, column=2, value="ORDER_SECTION").font = SUB_FONT
    ws.cell(row=3, column=2).fill = SECTION_FILL

    # Image 1: same_A rows 5-6 both
    set_row_heights(ws, 5, 6, ROW_H_IMAGE)
    ws.cell(row=5, column=2, value="IMG#1 same_A").font = MONO_FONT
    ws.cell(row=5, column=2).fill = OK_FILL
    add_image_twocell(ws, media["same_a"], row_start=5, row_end=6, col_start=3, col_end=5)

    # Image 2: same_B (left) / mod_b (right) rows 8-9
    set_row_heights(ws, 8, 9, ROW_H_IMAGE)
    if side == "left":
        ws.cell(row=8, column=2, value="IMG#2 same_B").font = MONO_FONT
        ws.cell(row=8, column=2).fill = OK_FILL
        add_image_twocell(ws, media["same_b"], row_start=8, row_end=9, col_start=3, col_end=5)
    else:
        ws.cell(row=8, column=2, value="IMG#2 mod_b").font = MONO_FONT
        ws.cell(row=8, column=2).fill = WARN_FILL
        add_image_twocell(ws, media["mod_b"], row_start=8, row_end=9, col_start=3, col_end=5)

    # Decoy on right only at different size — should NOT pair with left same_B
    # Place far below so order of primary images stays 1..2
    if side == "right":
        set_row_heights(ws, 12, 12, ROW_H_IMAGE)
        ws.cell(row=12, column=2, value="IMG decoy (誤ペア誘発用)").font = MONO_FONT
        ws.cell(row=12, column=2).fill = WARN_FILL
        add_image_twocell(ws, media["decoy"], row_start=12, row_end=12, col_start=3, col_end=6)
    else:
        # left-only small image
        set_row_heights(ws, 12, 12, ROW_H_IMAGE)
        ws.cell(row=12, column=2, value="IMG only_left_Y").font = MONO_FONT
        ws.cell(row=12, column=2).fill = WARN_FILL
        add_image_twocell(ws, media["only_left"], row_start=12, row_end=12, col_start=3, col_end=5)

    ws.cell(
        row=14,
        column=2,
        value="検証: #1 exact / #2 modified（同順）。decoy/only_left は片側のみで誤ペア禁止。",
    ).fill = NOTE_FILL
    ws.cell(row=14, column=2).font = Font(
        name="Yu Gothic UI", size=9, italic=True, color="666666"
    )


def build_cover(ws, side: str):
    """表紙（任意・同名で自動対応）。"""
    set_widths(ws, {"A": 4, "B": 28, "C": 52})
    ws["B2"] = "DiffXL 内容同期スクロール専用サンプル"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "左 (基準)" if side == "left" else "右 (比較対象)"
    ws["B3"].font = SUB_FONT

    info = [
        ("用途", "ContentScrollMap / 画像ギャップホールド / 再同期の検証"),
        ("版", "CS-L1.0" if side == "left" else "CS-R1.0"),
        ("対", "content_scroll_left.xlsx ↔ content_scroll_right.xlsx"),
        ("期待", "content_scroll_expected.json"),
    ]
    for i, (k, v) in enumerate(info):
        r = 5 + i
        ws.cell(row=r, column=2, value=k).font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="D9E2F3")
        ws.cell(row=r, column=2).border = THIN
        cell = ws.cell(row=r, column=3, value=v)
        cell.font = BODY_FONT
        cell.border = THIN

    ws["B10"] = "シート"
    ws["C10"] = "検証ポイント"
    style_header_row(ws, 10, 2, start_col=2)

    sheets = [
        ("SC_画像ギャップ", "右のみ区間ホールド → same_B 再同期"),
        ("SC_テキスト挿入", "2 行挿入後の S03 再連結"),
        ("SC_大画像span", "twoCell 複数行スパン"),
        ("SC_横同期", "列 1:1 横スクロール"),
        ("SC_同順異内容", "同順・異内容の誤ペア禁止"),
    ]
    for i, (a, b) in enumerate(sheets):
        r = 11 + i
        ws.cell(row=r, column=2, value=a).font = BODY_FONT
        ws.cell(row=r, column=3, value=b).font = BODY_FONT
        for c in range(2, 4):
            ws.cell(row=r, column=c).border = THIN


def create_workbook(side: str, media: dict[str, Path]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "表紙"
    build_cover(ws, side)

    ws = wb.create_sheet("SC_画像ギャップ")
    build_sc_image_gap(ws, side, media)

    ws = wb.create_sheet("SC_テキスト挿入")
    build_sc_text_insert(ws, side)

    ws = wb.create_sheet("SC_大画像span")
    build_sc_large_span(ws, side, media)

    ws = wb.create_sheet("SC_横同期")
    build_sc_horizontal(ws, side)

    ws = wb.create_sheet("SC_同順異内容")
    build_sc_same_order_diff(ws, side, media)

    return wb


def build_expected() -> dict:
    """content_scroll_expected.json スキーマ（brief Step 3 準拠 + 他シート補足）。"""
    return {
        "version": 1,
        "description": "DiffXL content-scroll perfect samples expected map",
        "files": {
            "left": "content_scroll_left.xlsx",
            "right": "content_scroll_right.xlsx",
        },
        "sheets": {
            "SC_画像ギャップ": {
                "imagePairs": [
                    {"leftRowStart": 5, "rightRowStart": 5, "kind": "exact"},
                    {"leftRowStart": 8, "rightRowStart": 12, "kind": "exact"},
                ],
                "leftOnly": [],
                "rightOnly": [{"rightRowStart": 8, "kind": "rightOnly"}],
                "scrollSamples": [
                    {"from": "L", "row": 5, "expectOther": 5},
                    {"from": "R", "row": 9, "expectOtherMax": 7},
                    {"from": "L", "row": 8, "expectOther": 12},
                    {"from": "R", "row": 12, "expectOther": 8},
                ],
            },
            "SC_テキスト挿入": {
                "textLandmarks": [
                    {"id": "S01", "leftRow": 5, "rightRow": 5},
                    {"id": "S02", "leftRow": 7, "rightRow": 7},
                    {"id": "S03", "leftRow": 10, "rightRow": 12},
                    {"id": "S04", "leftRow": 12, "rightRow": 14},
                    {"id": "S05", "leftRow": 14, "rightRow": 16},
                ],
                "scrollSamples": [
                    {
                        "from": "L",
                        "row": 10,
                        "expectOther": 12,
                        "note": "S03 after 2 insert rows",
                    },
                    {
                        "from": "R",
                        "row": 8,
                        "expectOtherMax": 7,
                        "note": "insert zone holds left",
                    },
                ],
            },
            "SC_大画像span": {
                "imagePairs": [
                    {
                        "leftRowStart": 5,
                        "leftRowEnd": 10,
                        "rightRowStart": 5,
                        "rightRowEnd": 10,
                        "kind": "exact",
                    },
                    {
                        "leftRowStart": 12,
                        "leftRowEnd": 13,
                        "rightRowStart": 12,
                        "rightRowEnd": 13,
                        "kind": "exact",
                    },
                ],
                "scrollSamples": [
                    {"from": "L", "row": 5, "expectOther": 5, "note": "span start"},
                    {"from": "L", "row": 8, "expectOther": 8, "note": "mid span holds pair"},
                    {"from": "L", "row": 12, "expectOther": 12, "note": "tail image"},
                ],
            },
            "SC_横同期": {
                "horizontal": {"mode": "oneToOne", "note": "columns always 1:1"},
                "scrollSamples": [
                    {"from": "L", "row": 5, "expectOther": 5},
                    {"from": "L", "row": 14, "expectOther": 14},
                    {"from": "L", "row": 24, "expectOther": 24},
                ],
            },
            "SC_同順異内容": {
                "imagePairs": [
                    {"leftRowStart": 5, "rightRowStart": 5, "kind": "exact"},
                    {"leftRowStart": 8, "rightRowStart": 8, "kind": "modified"},
                ],
                "leftOnly": [{"leftRowStart": 12, "kind": "leftOnly"}],
                "rightOnly": [{"rightRowStart": 12, "kind": "rightOnly"}],
                "scrollSamples": [
                    {"from": "L", "row": 5, "expectOther": 5},
                    {"from": "L", "row": 8, "expectOther": 8},
                ],
            },
        },
    }


def file_info(path: Path) -> str:
    return f"{path.name}  {path.stat().st_size:,} bytes"


def main():
    t0 = time.time()
    print("Generating content-scroll media...")
    media = generate_media()
    for k, p in media.items():
        print(f"  {k}: {p.name} ({p.stat().st_size:,} bytes)")

    left_path = OUT_DIR / "content_scroll_left.xlsx"
    right_path = OUT_DIR / "content_scroll_right.xlsx"
    expected_path = OUT_DIR / "content_scroll_expected.json"

    print("Building workbooks...")
    wb_l = create_workbook("left", media)
    wb_l.save(left_path)
    wb_r = create_workbook("right", media)
    wb_r.save(right_path)

    expected = build_expected()
    expected_path.write_text(
        json.dumps(expected, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print("\nOutputs:")
    print(" ", file_info(left_path))
    print(" ", file_info(right_path))
    print(" ", file_info(expected_path))
    print(f"\nDone in {time.time() - t0:.1f}s")
    print("Sheets: 表紙, SC_画像ギャップ, SC_テキスト挿入, SC_大画像span, SC_横同期, SC_同順異内容")


if __name__ == "__main__":
    main()
