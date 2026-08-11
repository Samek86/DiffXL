# -*- coding: utf-8 -*-
"""
DiffXL 機能検証用サンプル .xlsx を生成する。

出力先: 親ディレクトリ (30_参考資料/samples/)
  - full_feature_left.xlsx
  - full_feature_right.xlsx
  - README.md

検証観点:
  - テキスト差分 / 画像差分 / 片側のみ画像
  - 同名シート自動対応 / 別名シート手動対応
  - シート構成差分（左右どちらかのみのシート）
  - 行挿入による位置ずれ（アンカー設定の練習）
  - 長尺シート（同期スクロール・MiniMap）
  - フォント・行高・列幅・画像表示の忠実性
"""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT.parent
MEDIA = ROOT / "media"
MEDIA.mkdir(parents=True, exist_ok=True)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Yu Gothic UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Yu Gothic UI", size=18, bold=True, color="1F4E79")
SUB_FONT = Font(name="Yu Gothic UI", size=12, bold=True, color="2E75B6")
BODY_FONT = Font(name="Yu Gothic UI", size=11)
MONO_FONT = Font(name="Consolas", size=10)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
DIFF_HINT_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
ALT_ROW = PatternFill("solid", fgColor="D6EAF8")


def _font_large():
    try:
        return ImageFont.truetype("segoeui.ttf", 28)
    except OSError:
        try:
            return ImageFont.truetype("arial.ttf", 28)
        except OSError:
            return ImageFont.load_default()


def _font_small():
    try:
        return ImageFont.truetype("segoeui.ttf", 16)
    except OSError:
        try:
            return ImageFont.truetype("arial.ttf", 16)
        except OSError:
            return ImageFont.load_default()


def make_badge(path: Path, label: str, bg: tuple, accent: tuple, size=(320, 180)):
    """角丸風のサンプル画像を生成する。"""
    img = Image.new("RGB", size, bg)
    draw = ImageDraw.Draw(img)
    # 枠
    draw.rectangle([8, 8, size[0] - 9, size[1] - 9], outline=accent, width=4)
    # 上部バー
    draw.rectangle([8, 8, size[0] - 9, 48], fill=accent)
    # 中央アクセント円
    cx, cy = size[0] // 2, size[1] // 2 + 10
    r = 42
    draw.ellipse([cx - r, cy - r, cx + r, cy + r], fill=accent)
    draw.ellipse([cx - r + 12, cy - r + 12, cx + r - 12, cy + r - 12], fill=bg)
    # ラベル
    font = _font_large()
    bbox = draw.textbbox((0, 0), label, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text(((size[0] - tw) / 2, 12), label, fill=(255, 255, 255), font=font)
    # フッタ
    footer = "DiffXL sample media"
    f2 = _font_small()
    bbox2 = draw.textbbox((0, 0), footer, font=f2)
    tw2 = bbox2[2] - bbox2[0]
    draw.text(((size[0] - tw2) / 2, size[1] - 36), footer, fill=accent, font=f2)
    img.save(path, "PNG")
    return path


def make_modified_badge(src: Path, dst: Path, stamp: str = "MOD"):
    """既存画像に差分スタンプを重ねて「部分変更」画像を作る。"""
    img = Image.open(src).convert("RGB")
    draw = ImageDraw.Draw(img)
    w, h = img.size
    # 右下に赤い帯（OpenCV で検出される明確な差分）
    draw.rectangle([w // 2, h // 2, w - 12, h - 12], fill=(220, 50, 50))
    font = _font_large()
    bbox = draw.textbbox((0, 0), stamp, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((w // 2 + (w // 2 - 12 - tw) / 2, h // 2 + (h // 2 - 12 - th) / 2), stamp, fill=(255, 255, 255), font=font)
    # 左上に黄色い三角
    draw.polygon([(12, 56), (70, 56), (12, 114)], fill=(255, 200, 0))
    img.save(dst, "PNG")
    return dst


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN


def set_col_widths(ws, widths: dict):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def add_note(ws, cell_ref: str, text: str):
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")
    cell.fill = NOTE_FILL


def build_cover(ws, side: str, version: str, revision_note: str):
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 4, "B": 28, "C": 48, "D": 22})
    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 32
    ws["B2"] = "DiffXL 機能検証サンプル"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"{'左 (Left / 基準)' if side == 'left' else '右 (Right / 比較対象)'} ブック"
    ws["B3"].font = SUB_FONT

    labels = [
        ("B5", "ファイル役割", "C5", "比較の左ペイン" if side == "left" else "比較の右ペイン"),
        ("B6", "サンプル版", "C6", version),
        ("B7", "対象アプリ", "C7", "DiffXL (.xlsx 比較)"),
        ("B8", "想定用途", "C8", "テキスト／画像／シート構成／手動対応／アンカー等の一括確認"),
        ("B9", "改訂メモ", "C9", revision_note),
    ]
    for bl, bval, cl, cval in labels:
        ws[bl] = bval
        ws[bl].font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws[bl].fill = PatternFill("solid", fgColor="D9E2F3")
        ws[bl].border = THIN
        ws[cl] = cval
        ws[cl].font = BODY_FONT
        ws[cl].border = THIN
        if bl == "B6" and side == "right":
            ws[cl].fill = DIFF_HINT_FILL  # 意図的差分ヒント

    ws["B11"] = "シート一覧と検証ポイント"
    ws["B11"].font = SUB_FONT

    headers = ["シート名", "主な検証内容", "左右の違い"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=12, column=i, value=h)
    style_header_row(ws, 12, 4)
    # col A unused, style B-D
    for c in range(2, 5):
        ws.cell(row=12, column=c).fill = HEADER_FILL
        ws.cell(row=12, column=c).font = HEADER_FONT

    rows = [
        ("表紙", "起動後の表示・基本情報", "版番号・改訂メモが異なる"),
        ("売上サマリ", "テキスト差分（数値・文言）", "金額・担当・備考が一部変更"),
        ("製品カタログ", "埋め込み画像の比較", "同一／内容差分／片側のみ画像"),
        ("長い一覧", "同期スクロール・MiniMap", "散在する数セル差分"),
        ("レイアウト確認", "行高・列幅・フォント再現", "ほぼ同一（表示忠実性用）"),
        ("仕様メモ_*", "手動シート対応付け", "シート名が左右で異なる"),
        ("*_のみメモ", "シート構成差分", "片側にしか存在しない"),
        ("ずれ試験", "行挿入による比較ずれ／アンカー", "途中に行が追加されている"),
    ]
    for i, (a, b, c) in enumerate(rows):
        r = 13 + i
        ws.cell(row=r, column=2, value=a).font = BODY_FONT
        ws.cell(row=r, column=3, value=b).font = BODY_FONT
        ws.cell(row=r, column=4, value=c).font = BODY_FONT
        for col in range(2, 5):
            ws.cell(row=r, column=col).border = THIN
            if i % 2:
                ws.cell(row=r, column=col).fill = ALT_ROW

    add_note(ws, "B22", "※ オレンジ系ハイライトのセルは「意図的な差分」の目印です（比較結果そのものではありません）。")


def build_sales(ws, side: str):
    set_col_widths(ws, {"A": 8, "B": 16, "C": 18, "D": 12, "E": 14, "F": 22, "G": 18})
    ws.row_dimensions[1].height = 22
    ws["A1"] = "月次売上サマリ（テキスト差分用）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    headers = ["No", "年月", "部門", "数量", "売上(千円)", "担当", "備考"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 7)
    ws.row_dimensions[3].height = 20

    # base data (left)
    data = [
        (1, "2026-01", "東日本", 120, 3450, "佐藤", "計画通り"),
        (2, "2026-02", "東日本", 98, 2980, "佐藤", ""),
        (3, "2026-03", "西日本", 150, 4120, "鈴木", "キャンペーンあり"),
        (4, "2026-04", "西日本", 132, 3890, "鈴木", ""),
        (5, "2026-05", "中部", 88, 2210, "田中", "在庫調整"),
        (6, "2026-06", "中部", 110, 3050, "田中", ""),
        (7, "2026-07", "東日本", 140, 4500, "佐藤", "大型案件"),
        (8, "2026-08", "西日本", 95, 2760, "鈴木", "見込み"),
    ]

    if side == "right":
        # intentional text diffs
        data = [
            (1, "2026-01", "東日本", 120, 3450, "佐藤", "計画通り"),  # same
            (2, "2026-02", "東日本", 105, 3120, "佐藤", "修正計上"),  # qty/amount/note changed
            (3, "2026-03", "西日本", 150, 4120, "鈴木", "キャンペーンあり"),  # same
            (4, "2026-04", "西日本", 132, 4010, "高橋", ""),  # amount + owner
            (5, "2026-05", "中部", 88, 2210, "田中", "在庫調整"),  # same
            (6, "2026-06", "中部", 110, 3050, "田中", "遅延なし"),  # note changed
            (7, "2026-07", "東日本", 145, 4680, "佐藤", "大型案件（増額）"),  # multi-field
            (8, "2026-08", "西日本", 95, 2760, "鈴木", "確定"),  # note
        ]

    for r_idx, row in enumerate(data):
        excel_row = 4 + r_idx
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if c_idx in (4, 5):
                cell.alignment = Alignment(horizontal="right", vertical="center")
        if side == "right" and r_idx in (1, 3, 5, 6, 7):
            for c in range(1, 8):
                ws.cell(row=excel_row, column=c).fill = DIFF_HINT_FILL

    # totals
    total_row = 12
    ws.cell(row=total_row, column=3, value="合計").font = Font(name="Yu Gothic UI", size=11, bold=True)
    if side == "left":
        ws.cell(row=total_row, column=4, value=933).font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws.cell(row=total_row, column=5, value=26960).font = Font(name="Yu Gothic UI", size=11, bold=True)
    else:
        ws.cell(row=total_row, column=4, value=945).font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws.cell(row=total_row, column=5, value=27400).font = Font(name="Yu Gothic UI", size=11, bold=True)
        ws.cell(row=total_row, column=4).fill = DIFF_HINT_FILL
        ws.cell(row=total_row, column=5).fill = DIFF_HINT_FILL
    total_fill = PatternFill("solid", fgColor="DDEBF7")
    for c in range(1, 8):
        cell = ws.cell(row=total_row, column=c)
        cell.border = THIN
        # 右ブックで既に差分ヒント色が付いているセルは維持
        if side != "right" or c not in (4, 5):
            cell.fill = total_fill

    ws["A14"] = "空欄セル・同一セル混在"
    ws["A14"].font = SUB_FONT
    ws["A15"] = "コメントA"
    ws["B15"] = "共通テキスト：差分なし" if side == "left" else "共通テキスト：差分なし"
    ws["A16"] = "コメントB"
    ws["B16"] = "左側の文言です" if side == "left" else "右側では文言が変わっています"
    if side == "right":
        ws["B16"].fill = DIFF_HINT_FILL
    ws["A17"] = "コメントC"
    ws["B17"] = "" if side == "left" else "右側だけに追記"
    if side == "right":
        ws["B17"].fill = DIFF_HINT_FILL

    add_note(ws, "A19", "検証: 数量・金額・担当・備考・合計・B16/B17 のテキスト差分が検出されること。")


def build_catalog(ws, side: str, images: dict):
    """製品カタログ: 画像比較用。"""
    set_col_widths(ws, {"A": 4, "B": 22, "C": 38, "D": 12, "E": 40})
    ws.row_dimensions[1].height = 24
    ws["B1"] = "製品カタログ（埋め込み画像比較用）"
    ws["B1"].font = TITLE_FONT

    ws["B3"] = "画像ID"
    ws["C3"] = "説明"
    ws["D3"] = "状態"
    ws["E3"] = "プレビュー"
    style_header_row(ws, 3, 5)
    for c in range(2, 6):
        ws.cell(row=3, column=c).fill = HEADER_FILL
        ws.cell(row=3, column=c).font = HEADER_FONT

    # Row layout for images: each product needs tall rows
    products = [
        # id, desc, status_left, status_right, image_key_left, image_key_right
        ("IMG-A", "共通ロゴ（内容同一）", "同一", "同一", "same", "same"),
        ("IMG-B", "バナー（右だけ部分変更）", "基準", "内容差分", "base", "modified"),
        ("IMG-C", "左のみ追加アイコン", "左のみ", "—", "left_only", None),
        ("IMG-D", "右のみ追加アイコン", "—", "右のみ", None, "right_only"),
        ("IMG-E", "サムネ（同一・別ファイル名）", "同一", "同一", "thumb", "thumb"),
    ]

    row = 4
    for pid, desc, st_l, st_r, key_l, key_r in products:
        status = st_l if side == "left" else st_r
        key = key_l if side == "left" else key_r
        ws.row_dimensions[row].height = 100
        ws.cell(row=row, column=2, value=pid).font = Font(name="Consolas", size=12, bold=True)
        ws.cell(row=row, column=3, value=desc).font = BODY_FONT
        st_cell = ws.cell(row=row, column=4, value=status)
        st_cell.font = BODY_FONT
        st_cell.alignment = Alignment(horizontal="center", vertical="center")
        if status in ("内容差分", "左のみ", "右のみ"):
            st_cell.fill = WARN_FILL
        elif status == "同一":
            st_cell.fill = OK_FILL
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = THIN
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center", wrap_text=True)

        if key and key in images:
            img = XLImage(str(images[key]))
            # fit roughly into cell area
            img.width = 160
            img.height = 90
            ws.add_image(img, f"E{row}")
        row += 1

    add_note(
        ws,
        "B10",
        "検証: IMG-A/E は差分なし、IMG-B は画像内容差分、IMG-C は左のみ、IMG-D は右のみとして扱われること。"
        if side == "left"
        else "検証: 右ブックでは IMG-B に赤い MOD スタンプと黄色三角が入り、IMG-D が追加されています。",
    )
    ws["B11"] = "補足テキスト（画像間の再同期用アンカー文字列）"
    ws["B11"].font = SUB_FONT
    ws["B12"] = "ANCHOR_TEXT_CATALOG_SECTION"
    ws["B12"].font = MONO_FONT
    ws["C12"] = "この文字列は左右で同一です（比較再連結の手掛かり）"


def build_long_list(ws, side: str):
    set_col_widths(ws, {"A": 8, "B": 14, "C": 28, "D": 12, "E": 16, "F": 24})
    ws["A1"] = "長い一覧（同期スクロール・MiniMap 用）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = "上下に散在する差分で MiniMap のジャンプと同期スクロールを確認できます。"
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")

    headers = ["行", "コード", "品名", "単位", "単価", "メモ"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 6)

    # 80 data rows
    n = 80
    for i in range(1, n + 1):
        r = 4 + i
        code = f"P-{i:04d}"
        name = f"サンプル品目 {i:03d}"
        unit = "式" if i % 5 == 0 else "個"
        price = 1000 + (i * 37) % 9000
        memo = ""
        # intentional diffs on right at top / middle / bottom
        if side == "right":
            if i == 3:
                name = "サンプル品目 003（改訂）"
                memo = "名称変更"
            elif i == 40:
                price = price + 500
                memo = "単価改定"
            elif i == 41:
                memo = "右のみメモ"
            elif i == 75:
                name = "サンプル品目 075★"
                unit = "セット"
                memo = "下部差分"

        ws.cell(row=r, column=1, value=i).font = BODY_FONT
        ws.cell(row=r, column=2, value=code).font = MONO_FONT
        ws.cell(row=r, column=3, value=name).font = BODY_FONT
        ws.cell(row=r, column=4, value=unit).font = BODY_FONT
        ws.cell(row=r, column=5, value=price).font = BODY_FONT
        ws.cell(row=r, column=6, value=memo).font = BODY_FONT
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN
            if i % 2 == 0:
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F2F2F2")
            if side == "right" and i in (3, 40, 41, 75):
                ws.cell(row=r, column=c).fill = DIFF_HINT_FILL

    add_note(ws, "A86", "検証: 行3・40・41・75 付近の差分が MiniMap に現れ、クリックでジャンプできること。")


def build_layout(ws, side: str):
    """表示忠実性: 行高・列幅・フォント。内容はほぼ同一。"""
    set_col_widths(ws, {"A": 6, "B": 20, "C": 14, "D": 10, "E": 35, "F": 8})
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 45
    ws.row_dimensions[7].height = 15

    ws["B1"] = "レイアウト確認シート"
    ws["B1"].font = Font(name="Yu Gothic UI", size=20, bold=True, color="C00000")
    ws["B2"] = "行の高さ・列の幅・フォントを Excel 単独表示と並べて確認する用途"
    ws["B2"].font = Font(name="Yu Gothic UI", size=9, italic=True)

    ws["B4"] = "Yu Gothic UI 11pt"
    ws["B4"].font = Font(name="Yu Gothic UI", size=11)
    ws["C4"] = "Meiryo 14pt Bold"
    ws["C4"].font = Font(name="Meiryo", size=14, bold=True)
    ws["D4"] = "Consolas"
    ws["D4"].font = Font(name="Consolas", size=12, color="0070C0")
    ws["E4"] = "大きな行高のセル（高さ45）"
    ws["E4"].font = Font(name="Yu Gothic UI", size=16)
    ws["E4"].alignment = Alignment(vertical="center")

    ws["B5"] = "狭い列"
    ws["B5"].font = Font(name="Yu Gothic UI", size=10)
    ws["C5"] = "色付き"
    ws["C5"].font = Font(name="Yu Gothic UI", size=12, color="FFFFFF")
    ws["C5"].fill = PatternFill("solid", fgColor="548235")
    ws["D5"] = "右寄せ数字"
    ws["E5"] = 123456.78
    ws["E5"].number_format = '#,##0.00'
    ws["E5"].font = Font(name="Consolas", size=14)
    ws["E5"].alignment = Alignment(horizontal="right")

    ws["B7"] = "結合セル（B7:D8）"
    ws["B7"].font = Font(name="Yu Gothic UI", size=12, bold=True)
    ws["B7"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B7"].fill = PatternFill("solid", fgColor="DDEBF7")
    ws.merge_cells("B7:D8")
    ws.row_dimensions[8].height = 22

    # tiny intentional text diff for right so sheet isn't completely empty of diffs
    ws["B10"] = "表示確認メモ"
    ws["B10"].font = SUB_FONT
    if side == "left":
        ws["B11"] = "このシートはレイアウト再現が主目的です（内容差は最小）。"
    else:
        ws["B11"] = "このシートはレイアウト再現が主目的です（内容差は最小）。※右は末尾メモのみ差分"
        ws["B12"] = "right-only layout note"
        ws["B12"].fill = DIFF_HINT_FILL
        ws["B12"].font = BODY_FONT

    add_note(ws, "B14", "検証: 行高・列幅・フォント・結合セル・数値書式が Excel 本体表示どおりであること。")


def build_spec(ws, side: str):
    """別名シート: 手動対応用。"""
    set_col_widths(ws, {"A": 6, "B": 22, "C": 50})
    title = "仕様メモ（旧名称）" if side == "left" else "仕様メモ（新名称）"
    ws["B1"] = title
    ws["B1"].font = TITLE_FONT
    ws["B2"] = (
        "左ブックのシート名: 仕様メモ_旧 ／ 右ブック: 仕様メモ_新 → 自動では対応しないため手動対応付けで検証"
        if side == "left"
        else "シート名が異なるため、シート対応付けダイアログで「仕様メモ_旧」↔「仕様メモ_新」を手動設定してください。"
    )
    ws["B2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")
    ws.merge_cells("B2:C2")

    headers = ["項目", "内容"]
    ws["B4"] = headers[0]
    ws["C4"] = headers[1]
    style_header_row(ws, 4, 3)
    ws["B4"].fill = HEADER_FILL
    ws["C4"].fill = HEADER_FILL
    ws["B4"].font = HEADER_FONT
    ws["C4"].font = HEADER_FONT

    items = [
        ("対象形式", ".xlsx のみ"),
        ("比較単位", "シート単位（同名優先）"),
        ("画像比較", "OpenCV x64"),
        ("差分色既定", "黄色 / 不透明度 50%"),
        ("ビュー", "Excel 本体埋め込み"),
    ]
    if side == "right":
        items[3] = ("差分色既定", "黄色 / 不透明度 50%（設定で変更可）")  # minor text diff
        items.append(("追加項目", "右ブックのみの仕様メモ行"))

    for i, (k, v) in enumerate(items):
        r = 5 + i
        ws.cell(row=r, column=2, value=k).font = BODY_FONT
        ws.cell(row=r, column=3, value=v).font = BODY_FONT
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=3).border = THIN
        if side == "right" and i >= 3:
            ws.cell(row=r, column=2).fill = DIFF_HINT_FILL
            ws.cell(row=r, column=3).fill = DIFF_HINT_FILL

    add_note(ws, "B12", "検証: 手動シート対応後にテキスト差分が比較されること。")


def build_side_only(ws, side: str):
    set_col_widths(ws, {"A": 4, "B": 40, "C": 40})
    if side == "left":
        ws["B1"] = "左のみメモ"
        ws["B1"].font = TITLE_FONT
        ws["B3"] = "このシートは左ブックにのみ存在します。"
        ws["B4"] = "シート構成差分（Structure / 左のみ）として検出される想定です。"
        ws["B6"] = "内容サンプル"
        ws["B6"].font = SUB_FONT
        ws["B7"] = "左専用の設計メモ: 比較前チェックリスト"
        ws["B8"] = "・ファイルが .xlsx か"
        ws["B9"] = "・Excel がインストールされているか"
    else:
        ws["B1"] = "右のみメモ"
        ws["B1"].font = TITLE_FONT
        ws["B3"] = "このシートは右ブックにのみ存在します。"
        ws["B4"] = "シート構成差分（Structure / 右のみ）として検出される想定です。"
        ws["B6"] = "内容サンプル"
        ws["B6"].font = SUB_FONT
        ws["B7"] = "右専用のリリースメモ"
        ws["B8"] = "・差分強調トグル確認"
        ws["B9"] = "・設定 YAML の永続化確認"
    for cell in ("B3", "B4", "B7", "B8", "B9"):
        ws[cell].font = BODY_FONT
    add_note(ws, "B11", "検証: 比較結果にシート構成差分が出ること。手動対応の対象外シートとして扱う。")


def build_shift_test(ws, side: str):
    """途中行挿入で位置ずれ → アンカー設定の練習用。"""
    set_col_widths(ws, {"A": 8, "B": 20, "C": 36, "D": 16})
    ws["A1"] = "ずれ試験（行挿入・アンカー用）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        "左は連続した手順書。右は途中に挿入行があり、単純な行番号対応がずれます。"
        "アンカーを指定して再同期できるか確認してください。"
    )
    ws["A2"].font = Font(name="Yu Gothic UI", size=9, italic=True, color="666666")
    ws.merge_cells("A2:D2")

    headers = ["#", "ステップID", "作業内容", "担当"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 4)

    steps = [
        ("S01", "入力ファイルを用意する", "運用"),
        ("S02", "DiffXL を起動する", "運用"),
        ("S03", "左右の xlsx を選択する", "運用"),
        ("S04", "比較を実行する", "運用"),
        ("S05", "テキスト差分を確認する", "QA"),
        ("S06", "画像差分を確認する", "QA"),
        ("S07", "MiniMap でジャンプする", "QA"),
        ("S08", "差分強調を OFF/ON する", "QA"),
        ("S09", "設定で色を変える", "QA"),
        ("S10", "結果を記録する", "運用"),
    ]

    if side == "left":
        rows = list(steps)
    else:
        # insert two extra steps after S03, and tweak S07 text
        rows = steps[:3] + [
            ("S03a", "【挿入】シート対応を確認する", "運用"),
            ("S03b", "【挿入】必要なら手動対応する", "運用"),
        ] + steps[3:]
        # modify a later step text
        rows = list(rows)
        for i, (sid, text, owner) in enumerate(rows):
            if sid == "S07":
                rows[i] = (sid, "MiniMap でジャンプする（操作確認）", owner)

    for i, (sid, text, owner) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1).font = BODY_FONT
        ws.cell(row=r, column=2, value=sid).font = MONO_FONT
        ws.cell(row=r, column=3, value=text).font = BODY_FONT
        ws.cell(row=r, column=4, value=owner).font = BODY_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = THIN
            if side == "right" and sid in ("S03a", "S03b", "S07"):
                ws.cell(row=r, column=c).fill = DIFF_HINT_FILL

    # shared anchor text near bottom
    anchor_row = 5 + len(rows) + 2
    ws.cell(row=anchor_row, column=2, value="COMMON_ANCHOR_TOKEN").font = MONO_FONT
    ws.cell(row=anchor_row, column=3, value="左右同一のアンカー文字列（手動アンカーの目印）").font = BODY_FONT
    add_note(
        ws,
        f"A{anchor_row + 2}",
        "検証: 自動比較がずれる場合、アンカー設定で S04 以降や COMMON_ANCHOR_TOKEN を基準に再比較できること。",
    )


def create_workbook(side: str, images: dict) -> Workbook:
    wb = Workbook()
    # Cover
    ws = wb.active
    ws.title = "表紙"
    build_cover(
        ws,
        side,
        version="1.0.0" if side == "left" else "1.1.0",
        revision_note="初版（基準）" if side == "left" else "画像差分・行挿入・文言改訂を反映",
    )

    ws = wb.create_sheet("売上サマリ")
    build_sales(ws, side)

    ws = wb.create_sheet("製品カタログ")
    build_catalog(ws, side, images)

    ws = wb.create_sheet("長い一覧")
    build_long_list(ws, side)

    ws = wb.create_sheet("レイアウト確認")
    build_layout(ws, side)

    if side == "left":
        ws = wb.create_sheet("仕様メモ_旧")
        build_spec(ws, side)
        ws = wb.create_sheet("左のみメモ")
        build_side_only(ws, side)
    else:
        ws = wb.create_sheet("仕様メモ_新")
        build_spec(ws, side)
        ws = wb.create_sheet("右のみメモ")
        build_side_only(ws, side)

    ws = wb.create_sheet("ずれ試験")
    build_shift_test(ws, side)

    return wb


def write_readme():
    text = """# DiffXL サンプル Excel

このフォルダのサンプルは、DiffXL の主要機能を一通り確認するためのものです。

## 推奨ペア（フル機能）

| 左 (Left) | 右 (Right) |
|-----------|------------|
| `full_feature_left.xlsx` | `full_feature_right.xlsx` |

### シート対応と検証ポイント

| シート（左） | シート（右） | 自動対応 | 確認したいこと |
|--------------|--------------|----------|----------------|
| 表紙 | 表紙 | ○ 同名 | 基本表示、版番号テキスト差分 |
| 売上サマリ | 売上サマリ | ○ | **テキスト差分**（数量・金額・担当・備考・合計） |
| 製品カタログ | 製品カタログ | ○ | **画像差分**（同一 / 内容変更 / 左のみ / 右のみ） |
| 長い一覧 | 長い一覧 | ○ | **同期スクロール**・**MiniMap**（上下に散在する差分） |
| レイアウト確認 | レイアウト確認 | ○ | **表示忠実性**（行高・列幅・フォント・結合セル） |
| 仕様メモ_旧 | 仕様メモ_新 | × 別名 | **手動シート対応付け** → 再比較 |
| 左のみメモ | （なし） | — | **シート構成差分**（左のみ） |
| （なし） | 右のみメモ | — | **シート構成差分**（右のみ） |
| ずれ試験 | ずれ試験 | ○ | **行挿入によるずれ**・**アンカー設定** |

### 製品カタログの画像

| ID | 左 | 右 | 期待 |
|----|----|----|------|
| IMG-A | 共通ロゴ | 同一内容 | 画像差分なし |
| IMG-B | 基準バナー | 部分変更（赤 MOD + 黄三角） | 画像内容差分 |
| IMG-C | 左のみアイコン | なし | ImageOnlyLeft |
| IMG-D | なし | 右のみアイコン | ImageOnlyRight |
| IMG-E | サムネ | 同一内容 | 画像差分なし |

### その他の操作確認

- 差分強調トグル ON/OFF（再比較不要）
- 設定画面で差分色・不透明度変更
- 片側ファイル差し替え → 再比較
- 同一ファイルを左右に選んだ場合（差分ゼロに近い）

## 既存の簡易サンプル

| ファイル | 用途 |
|----------|------|
| `text_only_left.xlsx` / `text_only_right.xlsx` | テキストのみの最小差分（スモーク） |
| `_smoke_plan02.xlsx` | 単体スモーク用 |

## 再生成

```text
python 30_参考資料/samples/_gen/create_samples.py
```

生成物は `30_参考資料/samples/` 直下に上書き出力されます。
中間 PNG は `_gen/media/` に置かれます。
"""
    (OUT_DIR / "README.md").write_text(text, encoding="utf-8")


def main():
    # Generate images
    same = make_badge(MEDIA / "img_same.png", "LOGO-A", (30, 60, 110), (70, 140, 220))
    base = make_badge(MEDIA / "img_base.png", "BANNER-B", (40, 90, 70), (80, 170, 100))
    modified = make_modified_badge(base, MEDIA / "img_modified.png", "MOD")
    left_only = make_badge(MEDIA / "img_left_only.png", "LEFT-C", (120, 50, 40), (200, 90, 70), size=(280, 160))
    right_only = make_badge(MEDIA / "img_right_only.png", "RIGHT-D", (90, 50, 120), (170, 100, 210), size=(280, 160))
    thumb = make_badge(MEDIA / "img_thumb.png", "THUMB-E", (50, 50, 50), (180, 180, 180), size=(240, 140))

    images_left = {
        "same": same,
        "base": base,
        "left_only": left_only,
        "thumb": thumb,
    }
    images_right = {
        "same": same,
        "modified": modified,
        "right_only": right_only,
        "thumb": thumb,
    }

    left_path = OUT_DIR / "full_feature_left.xlsx"
    right_path = OUT_DIR / "full_feature_right.xlsx"

    wb_l = create_workbook("left", images_left)
    wb_l.save(left_path)

    wb_r = create_workbook("right", images_right)
    wb_r.save(right_path)

    write_readme()

    print("Created:")
    print(" ", left_path, left_path.stat().st_size, "bytes")
    print(" ", right_path, right_path.stat().st_size, "bytes")
    print(" ", OUT_DIR / "README.md")


if __name__ == "__main__":
    main()
